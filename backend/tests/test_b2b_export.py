from decimal import Decimal
from io import BytesIO
from zipfile import ZipFile

from httpx import AsyncClient
from openpyxl import load_workbook

from app.services.b2b_export import convert_to_kg, route_excel_filename
from tests.conftest import AuthUser, auth_header, login_token


def test_convert_to_kg_from_pieces() -> None:
    assert convert_to_kg(Decimal("0.25"), Decimal("400"), "шт") == Decimal("100.00")


def test_convert_to_kg_from_tons() -> None:
    assert convert_to_kg(Decimal("0.25"), Decimal("0.5"), "т") == Decimal("500.00")


def test_convert_to_kg_from_kg() -> None:
    assert convert_to_kg(Decimal("0.25"), Decimal("80"), "кг") == Decimal("80.00")


def test_route_excel_filename_uses_arrow() -> None:
    assert (
        route_excel_filename("Завод Московский", "Склад Ростов")
        == "Завод Московский → Склад Ростов.xlsx"
    )


async def test_export_b2b_forbidden_for_commercial(
    client: AsyncClient,
    test_user: AuthUser,
    catalog: dict,
) -> None:
    token = await login_token(client, test_user)
    response = await client.post(
        "/api/v1/logistics/normative/export-b2b",
        headers=auth_header(token),
        json={
            "routes": [
                {
                    "plant_code": catalog["plant_code"],
                    "warehouse_code": catalog["warehouse_code"],
                    "items": [
                        {
                            "product_code": catalog["product_code"],
                            "deficit": 400,
                            "unit": "шт",
                        }
                    ],
                }
            ]
        },
    )
    assert response.status_code == 403
    assert response.json()["error"]["code"] == "FORBIDDEN"


async def test_export_b2b_requires_routes(
    client: AsyncClient,
    logistics_user: AuthUser,
) -> None:
    token = await login_token(client, logistics_user)
    response = await client.post(
        "/api/v1/logistics/normative/export-b2b",
        headers=auth_header(token),
        json={"routes": []},
    )
    assert response.status_code == 400
    assert response.json()["error"]["code"] == "VALIDATION_ERROR"


async def test_export_b2b_unknown_product(
    client: AsyncClient,
    logistics_user: AuthUser,
    catalog: dict,
) -> None:
    token = await login_token(client, logistics_user)
    response = await client.post(
        "/api/v1/logistics/normative/export-b2b",
        headers=auth_header(token),
        json={
            "routes": [
                {
                    "plant_code": catalog["plant_code"],
                    "plant_name": "Завод Московский",
                    "warehouse_code": catalog["warehouse_code"],
                    "warehouse_name": "Склад Ростов",
                    "items": [
                        {"product_code": 99999999, "deficit": 10, "unit": "шт"}
                    ],
                }
            ]
        },
    )
    assert response.status_code == 400
    assert "99999999" in response.json()["error"]["message"]


async def test_export_b2b_returns_zip_with_kg(
    client: AsyncClient,
    logistics_user: AuthUser,
    catalog: dict,
) -> None:
    token = await login_token(client, logistics_user)
    response = await client.post(
        "/api/v1/logistics/normative/export-b2b",
        headers=auth_header(token),
        json={
            "routes": [
                {
                    "plant_code": catalog["plant_code"],
                    "plant_name": "Завод Московский",
                    "warehouse_code": catalog["warehouse_code"],
                    "warehouse_name": "Склад Ростов",
                    "items": [
                        {
                            "product_code": catalog["product_code"],
                            "product_name": "игнорируется",
                            "deficit": 400,
                            "unit": "шт",
                        }
                    ],
                },
                {
                    "plant_code": catalog["plant_code"],
                    "plant_name": "Завод Московский",
                    "warehouse_code": catalog["warehouse_code_2"],
                    "warehouse_name": "Склад Владивосток",
                    "items": [
                        {
                            "product_code": catalog["product_code_2"],
                            "deficit": 200,
                            "unit": "шт",
                        }
                    ],
                },
            ]
        },
    )
    assert response.status_code == 200, response.text
    assert response.headers["content-type"] == "application/zip"
    assert "b2b_orders_" in response.headers["content-disposition"]
    assert response.headers["content-disposition"].endswith('.zip"')

    with ZipFile(BytesIO(response.content)) as archive:
        names = archive.namelist()
        assert "Завод Московский → Склад Ростов.xlsx" in names
        assert "Завод Московский → Склад Владивосток.xlsx" in names

        workbook = load_workbook(BytesIO(archive.read(names[0])))
        sheet = workbook.active
        assert sheet.title == "Разнарядка"
        header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        assert header == ["Артикул", "Наименование", "Дефицит (кг)"]

        rostov = archive.read("Завод Московский → Склад Ростов.xlsx")
        sheet = load_workbook(BytesIO(rostov)).active
        row = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
        assert row[0] == catalog["product_code"]
        assert row[1] == "Подшипник 6204ZZ"
        assert row[2] == 100

        vladivostok = archive.read("Завод Московский → Склад Владивосток.xlsx")
        sheet = load_workbook(BytesIO(vladivostok)).active
        row = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
        assert row[0] == catalog["product_code_2"]
        assert row[1] == "Корпус чугунный 200мм"
        assert row[2] == 500


async def test_export_b2b_converts_tons_and_kg(
    client: AsyncClient,
    logistics_user: AuthUser,
    catalog: dict,
) -> None:
    token = await login_token(client, logistics_user)
    response = await client.post(
        "/api/v1/logistics/normative/export-b2b",
        headers=auth_header(token),
        json={
            "routes": [
                {
                    "plant_code": catalog["plant_code"],
                    "plant_name": "Завод Московский",
                    "warehouse_code": catalog["warehouse_code"],
                    "warehouse_name": "Склад Ростов",
                    "items": [
                        {
                            "product_code": catalog["product_code"],
                            "deficit": 0.5,
                            "unit": "т",
                        },
                        {
                            "product_code": catalog["product_code_2"],
                            "deficit": 12.5,
                            "unit": "кг",
                        },
                    ],
                }
            ]
        },
    )
    assert response.status_code == 200, response.text
    with ZipFile(BytesIO(response.content)) as archive:
        sheet = load_workbook(
            BytesIO(archive.read("Завод Московский → Склад Ростов.xlsx"))
        ).active
        rows = [
            [cell.value for cell in row]
            for row in sheet.iter_rows(min_row=2, max_row=3)
        ]
        by_code = {row[0]: row[2] for row in rows}
        assert by_code[catalog["product_code"]] == 500
        assert by_code[catalog["product_code_2"]] == 12.5
