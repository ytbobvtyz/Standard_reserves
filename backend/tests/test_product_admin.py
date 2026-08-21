from decimal import Decimal
from io import BytesIO

from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import delete

from app.core.database import AsyncSessionLocal
from app.models.available_balance import AvailableBalance
from app.models.product import Product
from tests.conftest import AuthUser, auth_header, login_token

NEW_CODE = 18201
UPDATE_CODE = 18202
DELETE_OK_CODE = 18203
DELETE_BLOCKED_CODE = 18204
DUP_GTIN_CODE = 18205
DUP_GTIN_SECOND = 18206
DANGLING_PARENT = 18999
DANGLING_CHILD = 18998


def _xlsx(rows: list[list[object | None]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


async def _cleanup_admin_products() -> None:
    codes = [
        NEW_CODE,
        UPDATE_CODE,
        DELETE_OK_CODE,
        DELETE_BLOCKED_CODE,
        DUP_GTIN_CODE,
        DUP_GTIN_SECOND,
        DANGLING_PARENT,
        DANGLING_CHILD,
    ]
    async with AsyncSessionLocal() as session:
        await session.execute(
            delete(AvailableBalance).where(AvailableBalance.product_code.in_(codes))
        )
        for code in codes:
            product = await session.get(Product, code)
            if product is not None:
                product.parent_code = None
                product.children_code = None
        await session.flush()
        await session.execute(delete(Product).where(Product.code.in_(codes)))
        await session.commit()


async def _seed_product(code: int, catalog: dict[str, int], **kwargs: object) -> None:
    async with AsyncSessionLocal() as session:
        session.add(
            Product(
                code=code,
                name=str(kwargs.get("name", f"Тестовый {code}")),
                category=str(kwargs.get("category", "A")),
                plant_id=catalog["plant_code"],
                weight_kg=Decimal("0.2500"),
                monthly_consumption=Decimal("100.00"),
                is_active=True,
                gtin=kwargs.get("gtin"),  # type: ignore[arg-type]
            )
        )
        await session.commit()


async def test_template_forbidden_for_commercial(
    client: AsyncClient, test_user: AuthUser, catalog: dict[str, int]
) -> None:
    token = await login_token(client, test_user)
    response = await client.get(
        "/api/v1/references/products/template",
        headers=auth_header(token),
    )
    assert response.status_code == 403


async def test_template_download_for_pp(
    client: AsyncClient, pp_user: AuthUser, catalog: dict[str, int]
) -> None:
    token = await login_token(client, pp_user)
    response = await client.get(
        "/api/v1/references/products/template",
        headers=auth_header(token),
    )
    assert response.status_code == 200, response.text
    assert "spreadsheetml" in response.headers["content-type"]
    workbook = load_workbook(BytesIO(response.content))
    headers = [cell.value for cell in next(workbook.active.iter_rows(max_row=1))]
    assert headers[:8] == [
        "code",
        "name",
        "status",
        "children_code",
        "parent_code",
        "category",
        "mark_control",
        "gtin",
    ]
    assert "plant_id" in headers
    assert "weight_kg" in headers


async def test_template_allowed_for_economist_and_logistics(
    client: AsyncClient,
    economist_user: AuthUser,
    logistics_user: AuthUser,
    catalog: dict[str, int],
) -> None:
    for user in (economist_user, logistics_user):
        token = await login_token(client, user)
        response = await client.get(
            "/api/v1/references/products/template",
            headers=auth_header(token),
        )
        assert response.status_code == 200, response.text


async def test_upload_creates_and_updates(
    client: AsyncClient, pp_user: AuthUser, catalog: dict[str, int]
) -> None:
    await _cleanup_admin_products()
    await _seed_product(UPDATE_CODE, catalog, name="Старое имя")
    token = await login_token(client, pp_user)
    content = _xlsx(
        [
            [
                "code",
                "name",
                "status",
                "children_code",
                "parent_code",
                "category",
                "mark_control",
                "gtin",
                "plant_id",
                "weight_kg",
            ],
            [
                NEW_CODE,
                "Новый артикул Excel",
                "Активный",
                None,
                None,
                "B",
                "да",
                "4601111111111",
                catalog["plant_code"],
                1.5,
            ],
            [
                UPDATE_CODE,
                "Обновлённое имя",
                "Активный",
                None,
                None,
                "A",
                "нет",
                "4602222222222",
                catalog["plant_code"],
                0.5,
            ],
            [
                "bad",
                "Ошибка",
                "Активный",
                None,
                None,
                "A",
                "нет",
                "123",
                catalog["plant_code"],
                1,
            ],
        ]
    )
    try:
        response = await client.post(
            "/api/v1/references/products/upload",
            headers=auth_header(token),
            files={
                "file": (
                    "products.xlsx",
                    content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200, response.text
        data = response.json()["data"]
        assert data["created"] == 1
        assert data["updated"] == 1
        assert data["errors"] == 1
        assert data["message"] == "Загружено 2, ошибок 1"

        created = await client.get(
            f"/api/v1/references/products/{NEW_CODE}/edit",
            headers=auth_header(token),
        )
        assert created.status_code == 200
        created_data = created.json()["data"]
        assert created_data["name"] == "Новый артикул Excel"
        assert created_data["gtin"] == "4601111111111"
        assert created_data["mark_control"] is True
        assert created_data["last_modified_by"]["id"] == str(pp_user.id)

        updated = await client.get(
            f"/api/v1/references/products/{UPDATE_CODE}",
            headers=auth_header(token),
        )
        assert updated.json()["data"]["name"] == "Обновлённое имя"
        assert updated.json()["data"]["gtin"] == "4602222222222"
    finally:
        await _cleanup_admin_products()


async def test_upload_allows_duplicate_gtin(
    client: AsyncClient, pp_user: AuthUser, catalog: dict[str, int]
) -> None:
    await _cleanup_admin_products()
    token = await login_token(client, pp_user)
    shared_gtin = "4605555555555"
    content = _xlsx(
        [
            [
                "code",
                "name",
                "status",
                "children_code",
                "parent_code",
                "category",
                "mark_control",
                "gtin",
                "plant_id",
                "weight_kg",
            ],
            [
                NEW_CODE,
                "Родительский артикул",
                "Активный",
                None,
                None,
                "A",
                "нет",
                shared_gtin,
                catalog["plant_code"],
                0.25,
            ],
            [
                DUP_GTIN_SECOND,
                "Дочерний артикул",
                "Активный",
                None,
                None,
                "B",
                "нет",
                shared_gtin,
                catalog["plant_code"],
                0.5,
            ],
        ]
    )
    try:
        response = await client.post(
            "/api/v1/references/products/upload",
            headers=auth_header(token),
            files={
                "file": (
                    "products.xlsx",
                    content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200, response.text
        data = response.json()["data"]
        assert data["created"] == 2
        assert data["errors"] == 0

        first = await client.get(
            f"/api/v1/references/products/{NEW_CODE}/edit",
            headers=auth_header(token),
        )
        second = await client.get(
            f"/api/v1/references/products/{DUP_GTIN_SECOND}/edit",
            headers=auth_header(token),
        )
        assert first.json()["data"]["gtin"] == shared_gtin
        assert second.json()["data"]["gtin"] == shared_gtin
    finally:
        await _cleanup_admin_products()


async def test_upload_allows_missing_parent_and_children(
    client: AsyncClient, pp_user: AuthUser, catalog: dict[str, int]
) -> None:
    await _cleanup_admin_products()
    token = await login_token(client, pp_user)
    content = _xlsx(
        [
            [
                "code",
                "name",
                "status",
                "children_code",
                "parent_code",
                "category",
                "mark_control",
                "gtin",
                "plant_id",
                "weight_kg",
            ],
            [
                NEW_CODE,
                "Артикул без связей в БД",
                "Активный",
                DANGLING_CHILD,
                DANGLING_PARENT,
                "C",
                "нет",
                "4606666666666",
                catalog["plant_code"],
                1.0,
            ],
        ]
    )
    try:
        response = await client.post(
            "/api/v1/references/products/upload",
            headers=auth_header(token),
            files={
                "file": (
                    "products.xlsx",
                    content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200, response.text
        data = response.json()["data"]
        assert data["created"] == 1
        assert data["errors"] == 0

        created = await client.get(
            f"/api/v1/references/products/{NEW_CODE}/edit",
            headers=auth_header(token),
        )
        assert created.status_code == 200
        created_data = created.json()["data"]
        assert created_data["parent_code"] == DANGLING_PARENT
        assert created_data["children_code"] == DANGLING_CHILD
        assert created_data["gtin"] == "4606666666666"
        assert created_data["category"] == "C"
    finally:
        await _cleanup_admin_products()


async def test_edit_update_delete_and_gtin_validation(
    client: AsyncClient, pp_user: AuthUser, catalog: dict[str, int]
) -> None:
    await _cleanup_admin_products()
    await _seed_product(DELETE_OK_CODE, catalog, name="На удаление")
    await _seed_product(DELETE_BLOCKED_CODE, catalog, name="Со связями")
    await _seed_product(DUP_GTIN_CODE, catalog, name="С GTIN", gtin="4603333333333")
    async with AsyncSessionLocal() as session:
        session.add(
            AvailableBalance(
                warehouse_code=catalog["warehouse_code"],
                product_code=DELETE_BLOCKED_CODE,
                available=Decimal("10"),
                plan=Decimal("10"),
                unit="шт",
            )
        )
        await session.commit()

    token = await login_token(client, pp_user)
    try:
        edit = await client.get(
            f"/api/v1/references/products/{DELETE_OK_CODE}/edit",
            headers=auth_header(token),
        )
        assert edit.status_code == 200
        body = edit.json()["data"]

        invalid = await client.put(
            f"/api/v1/references/products/{DELETE_OK_CODE}",
            headers=auth_header(token),
            json={
                **{
                    key: body[key]
                    for key in (
                        "name",
                        "description",
                        "category",
                        "is_active",
                        "weight_kg",
                        "monthly_consumption",
                        "mark_control",
                        "plant_id",
                        "second_plant_id",
                        "third_plant_id",
                        "parent_code",
                        "children_code",
                    )
                },
                "gtin": "12345",
            },
        )
        assert invalid.status_code == 400

        duplicate = await client.put(
            f"/api/v1/references/products/{DELETE_OK_CODE}",
            headers=auth_header(token),
            json={
                **{
                    key: body[key]
                    for key in (
                        "name",
                        "description",
                        "category",
                        "is_active",
                        "weight_kg",
                        "monthly_consumption",
                        "mark_control",
                        "plant_id",
                        "second_plant_id",
                        "third_plant_id",
                        "parent_code",
                        "children_code",
                    )
                },
                "gtin": "4603333333333",
            },
        )
        assert duplicate.status_code == 200, duplicate.text
        assert duplicate.json()["data"]["gtin"] == "4603333333333"

        updated = await client.put(
            f"/api/v1/references/products/{DELETE_OK_CODE}",
            headers=auth_header(token),
            json={
                "name": "Переименован",
                "description": None,
                "category": "C",
                "is_active": True,
                "weight_kg": 0.4,
                "monthly_consumption": 50,
                "gtin": "4604444444444",
                "mark_control": True,
                "plant_id": catalog["plant_code"],
                "second_plant_id": None,
                "third_plant_id": None,
                "parent_code": None,
                "children_code": None,
            },
        )
        assert updated.status_code == 200, updated.text
        data = updated.json()["data"]
        assert data["name"] == "Переименован"
        assert data["category"] == "C"
        assert data["gtin"] == "4604444444444"
        assert data["mark_control"] is True
        assert data["last_modified_by"]["full_name"] == pp_user.full_name

        blocked = await client.delete(
            f"/api/v1/references/products/{DELETE_BLOCKED_CODE}",
            headers=auth_header(token),
        )
        assert blocked.status_code == 409
        assert blocked.json()["error"]["code"] == "HAS_RELATIONS"

        deleted = await client.delete(
            f"/api/v1/references/products/{DELETE_OK_CODE}",
            headers=auth_header(token),
        )
        assert deleted.status_code == 200
        assert deleted.json()["message"] == "Продукт удален"

        missing = await client.get(
            f"/api/v1/references/products/{DELETE_OK_CODE}",
            headers=auth_header(token),
        )
        assert missing.status_code == 404
    finally:
        await _cleanup_admin_products()


async def test_guest_cannot_manage_products(
    client: AsyncClient, guest_user: AuthUser, catalog: dict[str, int]
) -> None:
    token = await login_token(client, guest_user)
    response = await client.get(
        "/api/v1/references/products/template",
        headers=auth_header(token),
    )
    assert response.status_code == 403
    listed = await client.get(
        "/api/v1/references/products",
        headers=auth_header(token),
    )
    assert listed.status_code == 200
