from datetime import date, timedelta
from io import BytesIO
from uuid import UUID

from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select

from app.core.database import AsyncSessionLocal
from app.models.normative import Normative
from app.models.production_request import ProductionRequest, ProductionRequestItem
from tests.conftest import AuthUser, auth_header, login_token


def _xlsx(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "Завод ERP",
            "Склад ERP",
            "Артикул",
            "Количество",
            "Ед.",
            "Клиент",
        ]
    )
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


async def test_production_request_template_starts_with_erp_columns(
    client: AsyncClient,
    logistics_user: AuthUser,
    catalog: dict[str, int],
) -> None:
    token = await login_token(client, logistics_user)
    response = await client.get(
        "/api/v1/production-requests/template",
        headers=auth_header(token),
    )
    assert response.status_code == 200, response.text
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    headers = [cell.value for cell in workbook.active[1]]
    assert headers[:3] == ["Завод ERP", "Склад ERP", "Артикул"]


async def test_upload_update_and_delete_production_request_batch(
    client: AsyncClient,
    logistics_user: AuthUser,
    catalog: dict[str, int],
) -> None:
    token = await login_token(client, logistics_user)
    valid_from = date.today() - timedelta(days=30)
    valid_to = date.today() + timedelta(days=90)
    content = _xlsx(
        [
            [
                catalog["erp_plant_code"],
                catalog["erp_warehouse_code"],
                catalog["product_code"],
                1000,
                "шт",
                "Клиент строки",
            ],
            [
                catalog["erp_plant_code_2"],
                catalog["erp_warehouse_code_2"],
                catalog["product_code_2"],
                500,
                "кг",
                None,
            ],
            [
                catalog["erp_plant_code"],
                catalog["erp_warehouse_code"],
                99999999,
                100,
                "шт",
                None,
            ],
            [
                catalog["erp_plant_code"],
                catalog["erp_warehouse_code"],
                catalog["product_code"],
                "не число",
                "шт",
                None,
            ],
        ]
    )
    response = await client.post(
        "/api/v1/production-requests/upload",
        headers=auth_header(token),
        files={
            "file": (
                "normatives.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "client_name": "Общий клиент",
            "valid_from": valid_from.isoformat(),
            "valid_to": valid_to.isoformat(),
        },
    )
    assert response.status_code == 201, response.text
    result = response.json()["data"]
    assert result["imported_count"] == 2
    assert result["total_rows"] == 4
    assert result["error_count"] == 2
    assert result["message"] == "Загружено 2 строк из 4"
    assert [item["row"] for item in result["error_details"]] == [4, 5]
    assert "артикул 99999999 не найден" in result["error_details"][0]["message"]
    assert "ожидается число" in result["error_details"][1]["message"]
    batch = result["production_request"]
    assert batch["source"] == "excel_upload"
    assert batch["items_count"] == 2
    assert batch["client_name"] == "Общий клиент"
    batch_id = UUID(batch["id"])

    try:
        async with AsyncSessionLocal() as session:
            items = (
                await session.scalars(
                    select(ProductionRequestItem).where(
                        ProductionRequestItem.production_request_id == batch_id
                    )
                )
            ).all()
            assert len(items) == 2
            item_ids = [item.id for item in items]
            normatives = (
                await session.scalars(
                    select(Normative).where(
                        Normative.production_request_item_id.in_(item_ids)
                    )
                )
            ).all()
            assert len(normatives) == 2
            assert all(item.request_id is None for item in normatives)
            assert {item.client_name for item in normatives} == {
                "Клиент строки",
                "Общий клиент",
            }
            assert {item.unit for item in normatives} == {"шт", "кг"}

        listing = await client.get(
            "/api/v1/production-requests",
            headers=auth_header(token),
        )
        assert listing.status_code == 200, listing.text
        listed_ids = {item["id"] for item in listing.json()["data"]}
        assert str(batch_id) in listed_ids

        new_from = date.today() - timedelta(days=60)
        new_to = date.today() + timedelta(days=180)
        updated = await client.patch(
            f"/api/v1/production-requests/{batch_id}/dates",
            headers=auth_header(token),
            json={
                "valid_from": new_from.isoformat(),
                "valid_to": new_to.isoformat(),
            },
        )
        assert updated.status_code == 200, updated.text
        assert updated.json()["data"]["valid_from"] == new_from.isoformat()
        assert updated.json()["data"]["valid_to"] == new_to.isoformat()

        async with AsyncSessionLocal() as session:
            expiries = (
                await session.scalars(
                    select(Normative.expiry_date).where(
                        Normative.production_request_item_id.in_(item_ids)
                    )
                )
            ).all()
            assert expiries == [new_to, new_to]

        deleted = await client.delete(
            f"/api/v1/production-requests/{batch_id}",
            headers=auth_header(token),
        )
        assert deleted.status_code == 200, deleted.text

        async with AsyncSessionLocal() as session:
            assert await session.get(ProductionRequest, batch_id) is None
            items_count = await session.scalar(
                select(func.count())
                .select_from(ProductionRequestItem)
                .where(ProductionRequestItem.production_request_id == batch_id)
            )
            normatives_count = await session.scalar(
                select(func.count())
                .select_from(Normative)
                .where(Normative.production_request_item_id.in_(item_ids))
            )
            assert items_count == 0
            assert normatives_count == 0
    finally:
        async with AsyncSessionLocal() as session:
            remaining = await session.get(ProductionRequest, batch_id)
            if remaining is not None:
                await session.delete(remaining)
                await session.commit()


async def test_commercial_cannot_manage_production_requests(
    client: AsyncClient,
    test_user: AuthUser,
    catalog: dict[str, int],
) -> None:
    token = await login_token(client, test_user)
    response = await client.get(
        "/api/v1/production-requests",
        headers=auth_header(token),
    )
    assert response.status_code == 403


async def test_economist_and_planner_can_list_production_requests(
    client: AsyncClient,
    economist_user: AuthUser,
    pp_user: AuthUser,
    catalog: dict[str, int],
) -> None:
    for user in (economist_user, pp_user):
        token = await login_token(client, user)
        response = await client.get(
            "/api/v1/production-requests",
            headers=auth_header(token),
        )
        assert response.status_code == 200, response.text
