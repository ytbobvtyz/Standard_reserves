from datetime import date
from decimal import Decimal
from io import BytesIO
from uuid import uuid4

import pytest
from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, select

from app.core.database import AsyncSessionLocal
from app.models.available_balance import AvailableBalance
from app.models.normative import Normative
from app.models.object import Object
from app.models.product import Product
from app.models.request import Request
from app.models.request_item import RequestItem
from app.models.sync_metadata import SyncMetadata
from app.services.coefficients import calculate_requirement
from app.services.logistics_normative import (
    LONG_DISTANCE_MESSAGE,
    _parse_int_cell,
    _parse_quantity,
    _parse_warehouse_code,
)
from tests.conftest import AuthUser, auth_header, delete_request, login_token

TEST_PRODUCT_DEFICIT = 19001
TEST_PRODUCT_OK = 19002
TEST_PRODUCT_STOCK_ONLY = 19003
CLIENT_NAME = "Тест логистика этап 5"


@pytest.fixture
async def logistics_catalog(catalog: dict[str, int], test_user: AuthUser):
    request_id = uuid4()
    warehouse_code = catalog["warehouse_code"]
    product_codes = [TEST_PRODUCT_DEFICIT, TEST_PRODUCT_OK, TEST_PRODUCT_STOCK_ONLY]
    async with AsyncSessionLocal() as session:
        await session.execute(
            delete(AvailableBalance).where(
                AvailableBalance.product_code.in_(product_codes)
            )
        )
        await session.execute(delete(SyncMetadata))
        await session.execute(
            delete(Normative).where(Normative.product_code.in_(product_codes))
        )
        leftover_ids = (
            await session.scalars(
                select(Request.id).where(Request.client_name == CLIENT_NAME)
            )
        ).all()
        await session.execute(
            delete(RequestItem).where(
                RequestItem.product_code.in_(product_codes)
                | RequestItem.request_id.in_(leftover_ids or [uuid4()])
            )
        )
        if leftover_ids:
            await session.execute(delete(Request).where(Request.id.in_(leftover_ids)))
        await session.execute(delete(Product).where(Product.code.in_(product_codes)))
        await session.commit()

    async with AsyncSessionLocal() as session:
        products = [
            Product(
                code=TEST_PRODUCT_DEFICIT,
                name="Тестовый подшипник логистики",
                category="A",
                plant_id=catalog["plant_code"],
                second_plant_id=1002,
                weight_kg=Decimal("0.2500"),
                is_active=True,
            ),
            Product(
                code=TEST_PRODUCT_OK,
                name="Тестовый корпус без дефицита",
                category="A",
                plant_id=1002,
                weight_kg=Decimal("2.5000"),
                is_active=True,
            ),
            Product(
                code=TEST_PRODUCT_STOCK_ONLY,
                name="Только фактический остаток",
                category="C",
                plant_id=catalog["plant_code"],
                weight_kg=Decimal("1.0000"),
                is_active=True,
            ),
        ]
        for product in products:
            if await session.get(Product, product.code) is None:
                session.add(product)
        await session.flush()
        session.add(
            Request(
                id=request_id,
                request_type="normative",
                status="active",
                client_name=CLIENT_NAME,
                initiator_id=test_user.id,
                expiry_date=date(2026, 12, 31),
            )
        )
        session.add(
            Normative(
                request_id=request_id,
                product_code=TEST_PRODUCT_DEFICIT,
                warehouse_code=warehouse_code,
                quantity=Decimal("1000"),
                unit="шт",
                client_name=CLIENT_NAME,
                expiry_date=date(2026, 12, 31),
                category="A",
            )
        )
        session.add(
            Normative(
                request_id=request_id,
                product_code=TEST_PRODUCT_OK,
                warehouse_code=warehouse_code,
                quantity=Decimal("500"),
                unit="шт",
                client_name=CLIENT_NAME,
                expiry_date=date(2026, 12, 31),
                category="A",
            )
        )
        session.add(
            AvailableBalance(
                warehouse_code=warehouse_code,
                product_code=TEST_PRODUCT_DEFICIT,
                available=Decimal("600"),
                plan=Decimal("600"),
                unit="шт",
                source="manual",
            )
        )
        session.add(
            AvailableBalance(
                warehouse_code=warehouse_code,
                product_code=TEST_PRODUCT_OK,
                available=Decimal("500"),
                plan=Decimal("500"),
                unit="шт",
                source="manual",
            )
        )
        session.add(
            AvailableBalance(
                warehouse_code=warehouse_code,
                product_code=TEST_PRODUCT_STOCK_ONLY,
                available=Decimal("80"),
                plan=Decimal("80"),
                unit="шт",
                source="manual",
            )
        )
        await session.commit()

    yield {
        "request_id": request_id,
        "warehouse_code": warehouse_code,
        "plant_code": catalog["plant_code"],
    }

    async with AsyncSessionLocal() as session:
        await session.execute(
            delete(AvailableBalance).where(
                AvailableBalance.product_code.in_(
                    [TEST_PRODUCT_DEFICIT, TEST_PRODUCT_OK, TEST_PRODUCT_STOCK_ONLY]
                )
            )
        )
        await session.commit()
    await delete_request(request_id)
    async with AsyncSessionLocal() as session:
        await session.execute(
            delete(Product).where(
                Product.code.in_(
                    [TEST_PRODUCT_DEFICIT, TEST_PRODUCT_OK, TEST_PRODUCT_STOCK_ONLY]
                )
            )
        )
        await session.commit()


def _warehouse(data: list[dict], warehouse_code: int) -> dict:
    return next(item for item in data if item["warehouse_code"] == warehouse_code)


def _item(warehouse: dict, product_code: int) -> dict:
    return next(
        item
        for item in warehouse["deficit_items"]
        if item["product_code"] == product_code
    )


async def test_dashboard_returns_deficit(
    client: AsyncClient,
    logistics_user: AuthUser,
    logistics_catalog: dict,
) -> None:
    token = await login_token(client, logistics_user)
    response = await client.get(
        "/api/v1/logistics/normative/dashboard",
        headers=auth_header(token),
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["status"] == "success"
    warehouse = _warehouse(body["data"], logistics_catalog["warehouse_code"])
    item = _item(warehouse, TEST_PRODUCT_DEFICIT)
    assert item["product_name"] == "Тестовый подшипник логистики"
    assert item["normative_quantity"] == 1000
    assert item["requirement"] == 1000
    assert item["available"] == 600
    assert item["plan"] == 600
    assert item["deficit"] == 400
    assert item["unit"] == "шт"
    assert item["status"] == "warning"
    assert item["client_name"] == CLIENT_NAME
    assert warehouse["long_distance"] is False
    assert warehouse["long_distance_message"] is None
    assert warehouse["deficit_count"] >= 1
    assert body["summary"]["deficit_products"] >= 1


def test_calculate_requirement_factors() -> None:
    assert calculate_requirement(Decimal("1000"), "A", False) == Decimal("1000")
    assert calculate_requirement(Decimal("1000"), "B", False) == Decimal("1500")
    assert calculate_requirement(Decimal("1000"), "C", False) == Decimal("2000")
    assert calculate_requirement(Decimal("1000"), "A", True) == Decimal("1500")
    assert calculate_requirement(Decimal("1000"), "B", True) == Decimal("2250")
    assert calculate_requirement(Decimal("1000"), "C", True) == Decimal("3000")


async def test_dashboard_requirement_uses_category_and_distance(
    client: AsyncClient,
    logistics_user: AuthUser,
    logistics_catalog: dict,
) -> None:
    warehouse_code = logistics_catalog["warehouse_code"]
    async with AsyncSessionLocal() as session:
        warehouse = await session.get(Object, warehouse_code)
        assert warehouse is not None
        warehouse.long_distance = True
        await session.commit()
    try:
        token = await login_token(client, logistics_user)
        response = await client.get(
            "/api/v1/logistics/normative/dashboard",
            params={"warehouse_code": warehouse_code},
            headers=auth_header(token),
        )
        assert response.status_code == 200, response.text
        warehouse = _warehouse(response.json()["data"], warehouse_code)
        assert warehouse["long_distance"] is True
        assert warehouse["long_distance_message"] == LONG_DISTANCE_MESSAGE
        item = _item(warehouse, TEST_PRODUCT_DEFICIT)
        assert item["normative_quantity"] == 1000
        assert item["requirement"] == 1500
        assert item["plan"] == 600
        assert item["deficit"] == 900
        assert item["category"] == "A"
    finally:
        async with AsyncSessionLocal() as session:
            warehouse = await session.get(Object, warehouse_code)
            assert warehouse is not None
            warehouse.long_distance = False
            await session.commit()


async def test_dashboard_remote_requirement_multiplies_category_and_distance(
    client: AsyncClient,
    logistics_user: AuthUser,
    logistics_catalog: dict,
) -> None:
    warehouse_code = logistics_catalog["warehouse_code"]
    async with AsyncSessionLocal() as session:
        warehouse = await session.get(Object, warehouse_code)
        assert warehouse is not None
        warehouse.long_distance = True
        ok_product = await session.get(Product, TEST_PRODUCT_OK)
        assert ok_product is not None
        ok_product.category = "B"
        await session.commit()
    try:
        token = await login_token(client, logistics_user)
        response = await client.get(
            "/api/v1/logistics/normative/dashboard",
            params={"warehouse_code": warehouse_code, "filter_mode": "all"},
            headers=auth_header(token),
        )
        assert response.status_code == 200, response.text
        warehouse = _warehouse(response.json()["data"], warehouse_code)
        assert warehouse["long_distance"] is True
        category_a = _item(warehouse, TEST_PRODUCT_DEFICIT)
        assert category_a["category"] == "A"
        assert category_a["normative_quantity"] == 1000
        assert category_a["requirement"] == 1500
        category_b = _item(warehouse, TEST_PRODUCT_OK)
        assert category_b["category"] == "B"
        assert category_b["normative_quantity"] == 500
        assert category_b["requirement"] == 1125
        assert category_b["plan"] == 500
        assert category_b["deficit"] == 625
    finally:
        async with AsyncSessionLocal() as session:
            warehouse = await session.get(Object, warehouse_code)
            assert warehouse is not None
            warehouse.long_distance = False
            ok_product = await session.get(Product, TEST_PRODUCT_OK)
            assert ok_product is not None
            ok_product.category = "A"
            await session.commit()


async def test_dashboard_filter_deficit_only(
    client: AsyncClient,
    logistics_user: AuthUser,
    logistics_catalog: dict,
) -> None:
    token = await login_token(client, logistics_user)
    response = await client.get(
        "/api/v1/logistics/normative/dashboard",
        params={"filter_mode": "deficit_only"},
        headers=auth_header(token),
    )
    assert response.status_code == 200, response.text
    warehouse = _warehouse(response.json()["data"], logistics_catalog["warehouse_code"])
    codes = {item["product_code"] for item in warehouse["deficit_items"]}
    assert TEST_PRODUCT_DEFICIT in codes
    assert TEST_PRODUCT_OK not in codes
    assert all(
        item["deficit"] > 0 or item.get("hide_group_metrics")
        for item in warehouse["deficit_items"]
    )


async def test_dashboard_filter_all_includes_stock_without_normative(
    client: AsyncClient,
    logistics_user: AuthUser,
    logistics_catalog: dict,
) -> None:
    token = await login_token(client, logistics_user)
    response = await client.get(
        "/api/v1/logistics/normative/dashboard",
        params={
            "filter_mode": "all",
            "warehouse_code": logistics_catalog["warehouse_code"],
        },
        headers=auth_header(token),
    )
    assert response.status_code == 200, response.text
    warehouse = _warehouse(response.json()["data"], logistics_catalog["warehouse_code"])
    codes = {item["product_code"] for item in warehouse["deficit_items"]}
    assert TEST_PRODUCT_STOCK_ONLY in codes
    stock_only = _item(warehouse, TEST_PRODUCT_STOCK_ONLY)
    assert stock_only["normative_quantity"] == 0
    assert stock_only["available"] == 80
    assert stock_only["plan"] == 80
    assert stock_only["status"] == "ok"


async def test_dashboard_unit_conversion(
    client: AsyncClient,
    logistics_user: AuthUser,
    logistics_catalog: dict,
) -> None:
    token = await login_token(client, logistics_user)
    response = await client.get(
        "/api/v1/logistics/normative/dashboard",
        params={"unit": "т", "warehouse_code": logistics_catalog["warehouse_code"]},
        headers=auth_header(token),
    )
    assert response.status_code == 200, response.text
    warehouse = _warehouse(response.json()["data"], logistics_catalog["warehouse_code"])
    item = _item(warehouse, TEST_PRODUCT_DEFICIT)
    assert item["unit"] == "т"
    assert item["normative_quantity"] == pytest.approx(0.25)
    assert item["requirement"] == pytest.approx(0.25)
    assert item["available"] == pytest.approx(0.15)
    assert item["plan"] == pytest.approx(0.15)
    assert item["deficit"] == pytest.approx(0.1)


async def test_dashboard_allowed_for_commercial(
    client: AsyncClient,
    test_user: AuthUser,
    logistics_catalog: dict,
) -> None:
    token = await login_token(client, test_user)
    response = await client.get(
        "/api/v1/logistics/normative/dashboard",
        headers=auth_header(token),
    )
    assert response.status_code == 200, response.text
    assert response.json()["status"] == "success"


async def test_generate_orders_forbidden_for_commercial(
    client: AsyncClient,
    test_user: AuthUser,
    logistics_catalog: dict,
) -> None:
    token = await login_token(client, test_user)
    warehouse_code = logistics_catalog["warehouse_code"]
    response = await client.post(
        f"/api/v1/logistics/normative/{warehouse_code}/generate-orders",
        headers=auth_header(token),
    )
    assert response.status_code == 403
    assert response.json()["error"]["code"] == "FORBIDDEN"


async def test_generate_orders_returns_correct_structure(
    client: AsyncClient,
    logistics_user: AuthUser,
    logistics_catalog: dict,
) -> None:
    token = await login_token(client, logistics_user)
    warehouse_code = logistics_catalog["warehouse_code"]
    response = await client.post(
        f"/api/v1/logistics/normative/{warehouse_code}/generate-orders",
        headers=auth_header(token),
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["total_orders"] >= 1
    assert data["total_products"] >= 1
    order = next(
        item
        for item in data["orders"]
        if item["plant_code"] == logistics_catalog["plant_code"]
    )
    assert order["warehouse_code"] == warehouse_code
    assert order["warehouse_name"] == "Склад Ростов"
    assert order["estimated_delivery_days"] == 5
    product = next(
        item for item in order["items"] if item["product_code"] == TEST_PRODUCT_DEFICIT
    )
    assert product["product_name"] == "Тестовый подшипник логистики"
    assert product["deficit"] == 400
    assert product["unit"] == "шт"
    assert product["weight_kg"] == pytest.approx(0.25)
    ok_codes = {
        item["product_code"] for order in data["orders"] for item in order["items"]
    }
    assert TEST_PRODUCT_OK not in ok_codes


async def test_generate_orders_unknown_warehouse(
    client: AsyncClient,
    logistics_user: AuthUser,
    logistics_catalog: dict,
) -> None:
    token = await login_token(client, logistics_user)
    response = await client.post(
        "/api/v1/logistics/normative/9999/generate-orders",
        headers=auth_header(token),
    )
    assert response.status_code == 404


async def test_generate_orders_bulk_returns_selected_warehouses(
    client: AsyncClient,
    logistics_user: AuthUser,
    logistics_catalog: dict,
) -> None:
    token = await login_token(client, logistics_user)
    warehouse_code = logistics_catalog["warehouse_code"]
    response = await client.post(
        "/api/v1/logistics/normative/generate-orders",
        headers=auth_header(token),
        json={"warehouse_codes": [warehouse_code, 2002]},
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["total_orders"] >= 1
    assert {order["warehouse_code"] for order in data["orders"]} <= {
        warehouse_code,
        2002,
    }
    product = next(
        item
        for order in data["orders"]
        for item in order["items"]
        if item["product_code"] == TEST_PRODUCT_DEFICIT
    )
    assert product["deficit"] == 400


async def test_generate_orders_bulk_requires_warehouse_codes(
    client: AsyncClient,
    logistics_user: AuthUser,
    logistics_catalog: dict,
) -> None:
    token = await login_token(client, logistics_user)
    empty = await client.post(
        "/api/v1/logistics/normative/generate-orders",
        headers=auth_header(token),
        json={"warehouse_codes": []},
    )
    assert empty.status_code == 400
    missing = await client.post(
        "/api/v1/logistics/normative/generate-orders",
        headers=auth_header(token),
        json={},
    )
    assert missing.status_code == 400


async def test_generate_orders_bulk_unknown_warehouse(
    client: AsyncClient,
    logistics_user: AuthUser,
    logistics_catalog: dict,
) -> None:
    token = await login_token(client, logistics_user)
    response = await client.post(
        "/api/v1/logistics/normative/generate-orders",
        headers=auth_header(token),
        json={"warehouse_codes": [9999]},
    )
    assert response.status_code == 404


async def test_export_excel_returns_file(
    client: AsyncClient,
    logistics_user: AuthUser,
    logistics_catalog: dict,
) -> None:
    token = await login_token(client, logistics_user)
    response = await client.get(
        "/api/v1/logistics/normative/export",
        params={"warehouse_code": logistics_catalog["warehouse_code"]},
        headers=auth_header(token),
    )
    assert response.status_code == 200, response.text
    assert "spreadsheetml.sheet" in response.headers["content-type"]
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert header == [
        "Склад",
        "Артикул",
        "Название",
        "Норматив",
        "Потребность",
        "Доступно",
        "Доступно + Запланировано",
        "Дефицит",
        "Ед",
        "Клиент",
    ]
    rows = [
        [cell.value for cell in row]
        for row in sheet.iter_rows(min_row=2, values_only=False)
    ]
    matched = next(row for row in rows if row[1] == TEST_PRODUCT_DEFICIT)
    assert matched[0] == "Склад Ростов"
    assert matched[2] == "Тестовый подшипник логистики"
    assert matched[3] == 1000
    assert matched[4] == 1000
    assert matched[5] == 600
    assert matched[6] == 600
    assert matched[7] == 400
    assert matched[8] == "шт"
    assert matched[9] == CLIENT_NAME


def _balance_excel_row(
    *,
    product_code: int | str | None,
    plant: int | str | None,
    warehouse: str | None,
    available: float | str | None,
    plan: float | str | None,
    unit: str | None = None,
) -> list:
    row: list = [None] * 20
    row[0] = product_code
    row[2] = plant
    row[4] = warehouse
    row[16] = unit
    row[18] = available
    row[19] = plan
    return row


def _build_balances_xlsx(data_rows: list[list]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        _balance_excel_row(
            product_code="product_code",
            plant="erp_plant_code",
            warehouse="erp_warehouse_code",
            available="available",
            plan="plan",
        )
    )
    for row in data_rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


async def test_upload_balances_upserts_and_reports_errors(
    client: AsyncClient,
    logistics_user: AuthUser,
    logistics_catalog: dict,
) -> None:
    token = await login_token(client, logistics_user)
    content = _build_balances_xlsx(
        [
            _balance_excel_row(
                product_code=TEST_PRODUCT_DEFICIT,
                plant=2401,
                warehouse="F005",
                available=500,
                plan=700,
            ),
            _balance_excel_row(
                product_code=TEST_PRODUCT_STOCK_ONLY,
                plant=2401,
                warehouse="F005",
                available=10,
                plan=15,
            ),
            _balance_excel_row(
                product_code=TEST_PRODUCT_OK,
                plant=2401,
                warehouse="X999",
                available=1,
                plan=1,
            ),
        ]
    )
    response = await client.post(
        "/api/v1/logistics/normative/upload",
        headers=auth_header(token),
        files={
            "file": (
                "balances.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["uploaded"] == 2
    assert data["updated"] >= 1
    assert data["errors"] == 1
    assert data["message"] == "Загружено 2, ошибок 1"
    assert any("X999" in item["message"] for item in data["error_details"])

    dashboard = await client.get(
        "/api/v1/logistics/normative/dashboard",
        params={"warehouse_code": logistics_catalog["warehouse_code"]},
        headers=auth_header(token),
    )
    assert dashboard.status_code == 200, dashboard.text
    warehouse = _warehouse(
        dashboard.json()["data"], logistics_catalog["warehouse_code"]
    )
    item = _item(warehouse, TEST_PRODUCT_DEFICIT)
    assert item["available"] == 500
    assert item["plan"] == 700
    assert item["deficit"] == 300
    assert item["stock_unit"] == "ШТ"

    async with AsyncSessionLocal() as session:
        stored = await session.get(
            AvailableBalance,
            (logistics_catalog["warehouse_code"], TEST_PRODUCT_DEFICIT),
        )
        assert stored is not None
        assert stored.unit == "ШТ"
        missing = await session.get(
            AvailableBalance,
            (logistics_catalog["warehouse_code"], TEST_PRODUCT_OK),
        )
        assert missing is None


async def test_upload_balances_replaces_warehouse_and_keeps_others(
    client: AsyncClient,
    logistics_user: AuthUser,
    logistics_catalog: dict,
) -> None:
    token = await login_token(client, logistics_user)
    other_warehouse = 2002
    async with AsyncSessionLocal() as session:
        session.add(
            AvailableBalance(
                warehouse_code=other_warehouse,
                product_code=TEST_PRODUCT_DEFICIT,
                available=Decimal("90"),
                plan=Decimal("90"),
                unit="шт",
                source="manual",
            )
        )
        await session.commit()

    content = _build_balances_xlsx(
        [
            _balance_excel_row(
                product_code=TEST_PRODUCT_DEFICIT,
                plant=2401,
                warehouse="F005",
                available=111,
                plan=222,
            )
        ]
    )
    response = await client.post(
        "/api/v1/logistics/normative/upload",
        headers=auth_header(token),
        files={
            "file": (
                "balances.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["uploaded"] == 1
    assert data["errors"] == 0

    async with AsyncSessionLocal() as session:
        rewritten = await session.get(
            AvailableBalance,
            (logistics_catalog["warehouse_code"], TEST_PRODUCT_DEFICIT),
        )
        gone_ok = await session.get(
            AvailableBalance,
            (logistics_catalog["warehouse_code"], TEST_PRODUCT_OK),
        )
        gone_stock = await session.get(
            AvailableBalance,
            (logistics_catalog["warehouse_code"], TEST_PRODUCT_STOCK_ONLY),
        )
        kept_other = await session.get(
            AvailableBalance,
            (other_warehouse, TEST_PRODUCT_DEFICIT),
        )
        assert rewritten is not None
        assert rewritten.available == Decimal("111")
        assert rewritten.plan == Decimal("222")
        assert rewritten.source == "excel"
        assert gone_ok is None
        assert gone_stock is None
        assert kept_other is not None
        assert kept_other.available == Decimal("90")
        await session.execute(
            delete(AvailableBalance).where(
                AvailableBalance.warehouse_code == other_warehouse,
                AvailableBalance.product_code == TEST_PRODUCT_DEFICIT,
            )
        )
        await session.commit()


async def test_sync_info_empty_before_upload(
    client: AsyncClient,
    logistics_user: AuthUser,
    logistics_catalog: dict,
) -> None:
    token = await login_token(client, logistics_user)
    response = await client.get(
        "/api/v1/logistics/normative/sync-info",
        headers=auth_header(token),
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["last_balances_sync_at"] is None
    assert data["last_balances_sync_by"] is None


async def test_sync_info_after_upload(
    client: AsyncClient,
    logistics_user: AuthUser,
    logistics_catalog: dict,
) -> None:
    token = await login_token(client, logistics_user)
    content = _build_balances_xlsx(
        [
            _balance_excel_row(
                product_code=TEST_PRODUCT_DEFICIT,
                plant=2401,
                warehouse="F005",
                available=1,
                plan=1,
            )
        ]
    )
    uploaded = await client.post(
        "/api/v1/logistics/normative/upload",
        headers=auth_header(token),
        files={
            "file": (
                "balances.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert uploaded.status_code == 200, uploaded.text

    response = await client.get(
        "/api/v1/logistics/normative/sync-info",
        headers=auth_header(token),
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["last_balances_sync_at"] is not None
    assert data["last_balances_sync_by"]["id"] == str(logistics_user.id)
    assert data["last_balances_sync_by"]["full_name"] == logistics_user.full_name
    assert data["last_balances_sync_by"]["role"] == "logistics"


async def test_sync_info_allowed_for_commercial(
    client: AsyncClient,
    test_user: AuthUser,
    logistics_catalog: dict,
) -> None:
    token = await login_token(client, test_user)
    response = await client.get(
        "/api/v1/logistics/normative/sync-info",
        headers=auth_header(token),
    )
    assert response.status_code == 200, response.text
    assert response.json()["status"] == "success"


async def test_upload_balances_forbidden_for_commercial(
    client: AsyncClient,
    test_user: AuthUser,
    logistics_catalog: dict,
) -> None:
    token = await login_token(client, test_user)
    content = _build_balances_xlsx(
        [
            _balance_excel_row(
                product_code=TEST_PRODUCT_DEFICIT,
                plant=2401,
                warehouse="F005",
                available=1,
                plan=1,
            )
        ]
    )
    response = await client.post(
        "/api/v1/logistics/normative/upload",
        headers=auth_header(token),
        files={
            "file": (
                "balances.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 403
    assert response.json()["error"]["code"] == "FORBIDDEN"


async def test_upload_balances_rejects_non_excel(
    client: AsyncClient,
    logistics_user: AuthUser,
    logistics_catalog: dict,
) -> None:
    token = await login_token(client, logistics_user)
    response = await client.post(
        "/api/v1/logistics/normative/upload",
        headers=auth_header(token),
        files={"file": ("balances.csv", b"not-excel", "text/csv")},
    )
    assert response.status_code == 400
    assert response.json()["error"]["code"] == "INVALID_FILE"


async def test_upload_balances_reads_unit_sht_kg_and_rejects_invalid(
    client: AsyncClient,
    logistics_user: AuthUser,
    logistics_catalog: dict,
) -> None:
    token = await login_token(client, logistics_user)
    content = _build_balances_xlsx(
        [
            _balance_excel_row(
                product_code=TEST_PRODUCT_DEFICIT,
                plant=2401,
                warehouse="F005",
                available=150,
                plan=150,
                unit="КГ",
            ),
            _balance_excel_row(
                product_code=TEST_PRODUCT_STOCK_ONLY,
                plant=2401,
                warehouse="F005",
                available=10,
                plan=10,
                unit=None,
            ),
            _balance_excel_row(
                product_code=TEST_PRODUCT_OK,
                plant=2401,
                warehouse="F005",
                available=1,
                plan=1,
                unit="ТОНН",
            ),
        ]
    )
    response = await client.post(
        "/api/v1/logistics/normative/upload",
        headers=auth_header(token),
        files={
            "file": (
                "balances.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["uploaded"] == 2
    assert data["errors"] == 1
    assert any(
        item["row"] == 4 and "ШТ или КГ" in item["message"]
        for item in data["error_details"]
    )

    async with AsyncSessionLocal() as session:
        kg_row = await session.get(
            AvailableBalance,
            (logistics_catalog["warehouse_code"], TEST_PRODUCT_DEFICIT),
        )
        empty_row = await session.get(
            AvailableBalance,
            (logistics_catalog["warehouse_code"], TEST_PRODUCT_STOCK_ONLY),
        )
        assert kg_row is not None
        assert kg_row.unit == "КГ"
        assert kg_row.available == Decimal("150")
        assert empty_row is not None
        assert empty_row.unit == "ШТ"

    dashboard = await client.get(
        "/api/v1/logistics/normative/dashboard",
        params={
            "warehouse_code": logistics_catalog["warehouse_code"],
            "unit": "шт",
        },
        headers=auth_header(token),
    )
    assert dashboard.status_code == 200, dashboard.text
    warehouse = _warehouse(
        dashboard.json()["data"], logistics_catalog["warehouse_code"]
    )
    item = _item(warehouse, TEST_PRODUCT_DEFICIT)
    assert item["stock_unit"] == "КГ"
    assert item["available"] == 600
    assert item["plan"] == 600
    assert item["deficit"] == 400


def test_parse_excel_codes_normalizes_float_and_string() -> None:
    assert _parse_int_cell(2401, "завод") == 2401
    assert _parse_int_cell(2401.0, "завод") == 2401
    assert _parse_int_cell("2401.0", "завод") == 2401
    assert _parse_int_cell("2401", "завод") == 2401
    assert _parse_warehouse_code("F005") == "F005"
    assert _parse_warehouse_code("f005") == "F005"
    assert _parse_warehouse_code(2401.0) == "2401"
    assert _parse_quantity(-120.5, "available") == Decimal("-120.50")
    assert _parse_quantity("-80,00", "plan") == Decimal("-80.00")
    assert _parse_quantity(0, "available") == Decimal("0.00")


async def test_upload_balances_finds_erp_plant_code_on_warehouse(
    client: AsyncClient,
    logistics_user: AuthUser,
    logistics_catalog: dict,
) -> None:
    """erp_plant_code may live on warehouses when there is no type=plant row."""
    token = await login_token(client, logistics_user)
    warehouse_code = 2002
    erp_plant_only_on_warehouse = 2410
    previous_plant_code = None
    async with AsyncSessionLocal() as session:
        warehouse = await session.get(Object, warehouse_code)
        assert warehouse is not None
        previous_plant_code = warehouse.erp_plant_code
        warehouse.erp_plant_code = erp_plant_only_on_warehouse
        await session.commit()

    try:
        content = _build_balances_xlsx(
            [
                _balance_excel_row(
                    product_code=TEST_PRODUCT_DEFICIT,
                    plant=2410.0,
                    warehouse="F006",
                    available=33,
                    plan=44,
                )
            ]
        )
        response = await client.post(
            "/api/v1/logistics/normative/upload",
            headers=auth_header(token),
            files={
                "file": (
                    "balances.xlsx",
                    content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 200, response.text
        data = response.json()["data"]
        assert data["uploaded"] == 1
        assert data["errors"] == 0
        assert data["error_details"] == []

        async with AsyncSessionLocal() as session:
            stored = await session.get(
                AvailableBalance,
                (warehouse_code, TEST_PRODUCT_DEFICIT),
            )
            assert stored is not None
            assert stored.available == Decimal("33")
            assert stored.plan == Decimal("44")
            await session.execute(
                delete(AvailableBalance).where(
                    AvailableBalance.warehouse_code == warehouse_code,
                    AvailableBalance.product_code == TEST_PRODUCT_DEFICIT,
                )
            )
            await session.commit()
    finally:
        async with AsyncSessionLocal() as session:
            warehouse = await session.get(Object, warehouse_code)
            if warehouse is not None:
                warehouse.erp_plant_code = previous_plant_code
                await session.commit()


async def test_upload_balances_accepts_negative_available_and_plan(
    client: AsyncClient,
    logistics_user: AuthUser,
    logistics_catalog: dict,
) -> None:
    token = await login_token(client, logistics_user)
    content = _build_balances_xlsx(
        [
            _balance_excel_row(
                product_code=TEST_PRODUCT_DEFICIT,
                plant=2401,
                warehouse="F005",
                available=-80,
                plan=-30,
            )
        ]
    )
    response = await client.post(
        "/api/v1/logistics/normative/upload",
        headers=auth_header(token),
        files={
            "file": (
                "balances.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["uploaded"] == 1
    assert data["errors"] == 0

    async with AsyncSessionLocal() as session:
        stored = await session.get(
            AvailableBalance,
            (logistics_catalog["warehouse_code"], TEST_PRODUCT_DEFICIT),
        )
        assert stored is not None
        assert stored.available == Decimal("-80")
        assert stored.plan == Decimal("-30")

    dashboard = await client.get(
        "/api/v1/logistics/normative/dashboard",
        params={"warehouse_code": logistics_catalog["warehouse_code"]},
        headers=auth_header(token),
    )
    assert dashboard.status_code == 200, dashboard.text
    warehouse = _warehouse(
        dashboard.json()["data"], logistics_catalog["warehouse_code"]
    )
    item = _item(warehouse, TEST_PRODUCT_DEFICIT)
    assert item["available"] == -80
    assert item["plan"] == -30
    assert item["normative_quantity"] == 1000
    assert item["deficit"] == 1030
    assert item["status"] == "warning"
