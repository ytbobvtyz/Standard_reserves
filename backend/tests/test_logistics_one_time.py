from datetime import UTC, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from uuid import uuid4

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy import select

from app.core.database import AsyncSessionLocal
from app.models.request import Request
from app.models.request_item import RequestItem
from tests.conftest import AuthUser, auth_header, delete_request, login_token

CLIENT_A = "Тест логистика этап 6 Ромашка"
CLIENT_B = "Тест логистика этап 6 Тюльпан"


@pytest.fixture
async def one_time_catalog(
    catalog: dict[str, int],
    test_user: AuthUser,
    other_user: AuthUser,
):
    ids: list = []
    created_old = datetime.now(UTC) - timedelta(days=10)
    created_new = datetime.now(UTC)

    async with AsyncSessionLocal() as session:
        leftover_ids = (
            await session.scalars(
                select(Request.id).where(
                    Request.client_name.like("Тест логистика этап 6%")
                )
            )
        ).all()
    for leftover_id in leftover_ids:
        await delete_request(leftover_id)

    payloads = [
        {
            "status": "approved",
            "client_name": CLIENT_A,
            "initiator_id": test_user.id,
            "warehouse_code": catalog["warehouse_code"],
            "product_code": catalog["product_code"],
            "quantity": Decimal("200"),
            "created_at": created_new,
        },
        {
            "status": "executed",
            "client_name": CLIENT_B,
            "initiator_id": other_user.id,
            "warehouse_code": catalog["warehouse_code_2"],
            "product_code": catalog["product_code_2"],
            "quantity": Decimal("50"),
            "created_at": created_old,
        },
        {
            "status": "rejected",
            "client_name": CLIENT_A,
            "initiator_id": test_user.id,
            "warehouse_code": catalog["warehouse_code"],
            "product_code": catalog["product_code_2"],
            "quantity": Decimal("10"),
            "created_at": created_new,
        },
    ]

    async with AsyncSessionLocal() as session:
        for payload in payloads:
            request_id = uuid4()
            ids.append(request_id)
            request = Request(
                id=request_id,
                request_type="one_time",
                status=payload["status"],
                client_name=payload["client_name"],
                initiator_id=payload["initiator_id"],
                created_at=payload["created_at"],
            )
            session.add(request)
            await session.flush()
            session.add(
                RequestItem(
                    request_id=request_id,
                    product_code=payload["product_code"],
                    warehouse_code=payload["warehouse_code"],
                    quantity_requested=payload["quantity"],
                    quantity_approved=payload["quantity"],
                    unit="шт",
                )
            )
        await session.commit()

    yield {
        "approved_id": ids[0],
        "executed_id": ids[1],
        "rejected_id": ids[2],
        "warehouse_code": catalog["warehouse_code"],
        "warehouse_code_2": catalog["warehouse_code_2"],
        "initiator_id": test_user.id,
        "other_initiator_id": other_user.id,
        "created_old": created_old,
        "created_new": created_new,
    }

    for request_id in ids:
        await delete_request(request_id)


async def test_one_time_list_requires_logistics(
    client: AsyncClient,
    test_user: AuthUser,
    one_time_catalog: dict,
) -> None:
    token = await login_token(client, test_user)
    response = await client.get(
        "/api/v1/logistics/one-time/list",
        headers=auth_header(token),
    )
    assert response.status_code == 403


async def test_one_time_list_filters(
    client: AsyncClient,
    logistics_user: AuthUser,
    one_time_catalog: dict,
) -> None:
    token = await login_token(client, logistics_user)
    headers = auth_header(token)

    listing = await client.get("/api/v1/logistics/one-time/list", headers=headers)
    assert listing.status_code == 200, listing.text
    payload = listing.json()
    assert payload["status"] == "success"
    assert "meta" in payload
    ids = {item["id"] for item in payload["data"]}
    assert str(one_time_catalog["approved_id"]) in ids
    approved = next(
        item
        for item in payload["data"]
        if item["id"] == str(one_time_catalog["approved_id"])
    )
    assert approved["client_name"] == CLIENT_A
    assert approved["status"] == "approved"
    assert approved["items"][0]["warehouse_code"] == one_time_catalog["warehouse_code"]
    assert approved["items"][0]["product_name"]
    assert approved["initiator"]["id"] == str(one_time_catalog["initiator_id"])

    by_warehouse = await client.get(
        "/api/v1/logistics/one-time/list",
        params={"warehouse_code": one_time_catalog["warehouse_code"]},
        headers=headers,
    )
    assert by_warehouse.status_code == 200
    warehouse_ids = {item["id"] for item in by_warehouse.json()["data"]}
    assert str(one_time_catalog["approved_id"]) in warehouse_ids
    assert str(one_time_catalog["executed_id"]) not in warehouse_ids

    by_client = await client.get(
        "/api/v1/logistics/one-time/list",
        params={"client_name": "тюльпан"},
        headers=headers,
    )
    assert by_client.status_code == 200
    client_ids = {item["id"] for item in by_client.json()["data"]}
    assert str(one_time_catalog["executed_id"]) in client_ids
    assert str(one_time_catalog["approved_id"]) not in client_ids

    by_initiator = await client.get(
        "/api/v1/logistics/one-time/list",
        params={"initiator_id": str(one_time_catalog["other_initiator_id"])},
        headers=headers,
    )
    assert by_initiator.status_code == 200
    initiator_ids = {item["id"] for item in by_initiator.json()["data"]}
    assert str(one_time_catalog["executed_id"]) in initiator_ids
    assert str(one_time_catalog["approved_id"]) not in initiator_ids

    old_day = one_time_catalog["created_old"].date().isoformat()
    by_date = await client.get(
        "/api/v1/logistics/one-time/list",
        params={"from_date": old_day, "to_date": old_day},
        headers=headers,
    )
    assert by_date.status_code == 200
    date_ids = {item["id"] for item in by_date.json()["data"]}
    assert str(one_time_catalog["executed_id"]) in date_ids
    assert str(one_time_catalog["approved_id"]) not in date_ids

    by_status = await client.get(
        "/api/v1/logistics/one-time/list",
        params={"status": "approved"},
        headers=headers,
    )
    assert by_status.status_code == 200
    status_ids = {item["id"] for item in by_status.json()["data"]}
    assert str(one_time_catalog["approved_id"]) in status_ids
    assert str(one_time_catalog["executed_id"]) not in status_ids
    assert str(one_time_catalog["rejected_id"]) not in status_ids


async def test_one_time_execute(
    client: AsyncClient,
    logistics_user: AuthUser,
    one_time_catalog: dict,
) -> None:
    token = await login_token(client, logistics_user)
    request_id = one_time_catalog["approved_id"]
    response = await client.post(
        f"/api/v1/logistics/one-time/{request_id}/execute",
        headers=auth_header(token),
        json={"order_number": "РН-2026-08-20-001", "comment": "Отгрузка произведена"},
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["id"] == str(request_id)
    assert data["status"] == "executed"
    assert data["order_number"] == "РН-2026-08-20-001"
    assert data["executed_comment"] == "Отгрузка произведена"
    assert data["executed_at"]
    assert data["executed_by"] == str(logistics_user.id)

    detail = await client.get(
        f"/api/v1/requests/{request_id}",
        headers=auth_header(token),
    )
    assert detail.status_code == 200
    assert detail.json()["data"]["status"] == "executed"

    listing = await client.get(
        "/api/v1/logistics/one-time/list",
        params={"status": "executed"},
        headers=auth_header(token),
    )
    matched = next(
        item for item in listing.json()["data"] if item["id"] == str(request_id)
    )
    assert matched["order_number"] == "РН-2026-08-20-001"


async def test_one_time_cannot_execute_wrong_status(
    client: AsyncClient,
    logistics_user: AuthUser,
    one_time_catalog: dict,
) -> None:
    token = await login_token(client, logistics_user)
    response = await client.post(
        f"/api/v1/logistics/one-time/{one_time_catalog['rejected_id']}/execute",
        headers=auth_header(token),
        json={"order_number": "РН-2026-08-20-002", "comment": "Нельзя"},
    )
    assert response.status_code == 400
    assert response.json()["error"]["code"] == "INVALID_STATUS"

    already = await client.post(
        f"/api/v1/logistics/one-time/{one_time_catalog['executed_id']}/execute",
        headers=auth_header(token),
        json={"order_number": "РН-2026-08-20-003"},
    )
    assert already.status_code == 400


async def test_one_time_execute_requires_order_number(
    client: AsyncClient,
    logistics_user: AuthUser,
    one_time_catalog: dict,
) -> None:
    token = await login_token(client, logistics_user)
    response = await client.post(
        f"/api/v1/logistics/one-time/{one_time_catalog['approved_id']}/execute",
        headers=auth_header(token),
        json={"comment": "Без номера"},
    )
    assert response.status_code == 422


async def test_one_time_initiators_and_clients(
    client: AsyncClient,
    logistics_user: AuthUser,
    test_user: AuthUser,
    other_user: AuthUser,
    one_time_catalog: dict,
) -> None:
    token = await login_token(client, logistics_user)
    initiators = await client.get(
        "/api/v1/logistics/one-time/initiators",
        headers=auth_header(token),
    )
    assert initiators.status_code == 200, initiators.text
    initiator_ids = {item["id"] for item in initiators.json()["data"]}
    assert str(test_user.id) in initiator_ids
    assert str(other_user.id) in initiator_ids
    assert all("full_name" in item for item in initiators.json()["data"])

    clients = await client.get(
        "/api/v1/logistics/one-time/clients",
        headers=auth_header(token),
    )
    assert clients.status_code == 200, clients.text
    names = set(clients.json()["data"])
    assert CLIENT_A in names
    assert CLIENT_B in names


async def test_one_time_export_excel(
    client: AsyncClient,
    logistics_user: AuthUser,
    one_time_catalog: dict,
) -> None:
    token = await login_token(client, logistics_user)
    request_id = one_time_catalog["approved_id"]
    response = await client.get(
        f"/api/v1/logistics/one-time/{request_id}/export",
        headers=auth_header(token),
    )
    assert response.status_code == 200
    assert "spreadsheetml.sheet" in response.headers["content-type"]
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert header == [
        "Артикул",
        "Название",
        "Склад",
        "Количество",
        "Ед",
        "Клиент",
        "Заявитель",
        "Статус",
    ]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert rows
    assert rows[0][5] == CLIENT_A
    assert rows[0][3] == 200
