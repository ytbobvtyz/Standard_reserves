from __future__ import annotations

import re
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import exists, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.exceptions import APIError
from app.models.available_balance import AvailableBalance
from app.models.normative import Normative
from app.models.object import Object
from app.models.product import Product
from app.models.request_item import RequestItem
from app.models.user import User
from app.schemas.reference import (
    ProductDetail,
    ProductUpdate,
    ProductUploadError,
    ProductUploadResult,
)
from app.services.audit import add_audit_log
from app.services.references import to_product_detail

GTIN_RE = re.compile(r"^[0-9]{13}$")
TEMPLATE_HEADERS = [
    "code",
    "name",
    "status",
    "children_code",
    "parent_code",
    "category",
    "mark_control",
    "gtin",
    "plant_id",
    "weight_kg",
]
EXAMPLE_ROW = [
    10020,
    "Подшипник 6205ZZ",
    "Активный",
    None,
    None,
    "A",
    "нет",
    "4601234567890",
    1001,
    0.25,
]


def build_template_xlsx() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "products"
    sheet.append(TEMPLATE_HEADERS)
    sheet.append(EXAMPLE_ROW)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _normalize_gtin(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in {"none", "nan"}:
        return None
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        text = text[:-2]
    if not GTIN_RE.fullmatch(text):
        raise ValueError("GTIN должен содержать 13 цифр")
    return text


def _parse_bool_mark(value: Any) -> bool:
    if value is None or str(value).strip() == "":
        return False
    text = str(value).strip().lower()
    if text in {"да", "yes", "true", "1"}:
        return True
    if text in {"нет", "no", "false", "0"}:
        return False
    raise ValueError("mark_control: укажите «да» или «нет»")


def _parse_active(value: Any) -> bool:
    if value is None or str(value).strip() == "":
        return False
    return str(value).strip().lower() == "активный"


def _parse_int(value: Any, field: str) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        raise ValueError(f"{field} должен быть числом") from None


def _parse_decimal(value: Any, field: str) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return Decimal(str(value).strip().replace(",", "."))
    except (InvalidOperation, TypeError, ValueError):
        raise ValueError(f"{field} должен быть числом") from None


def _cell_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, raw in enumerate(header_row):
        if raw is None:
            continue
        mapping[str(raw).strip().lower()] = index
    return mapping


async def _load_product(db: AsyncSession, code: int) -> Product | None:
    result = await db.execute(
        select(Product)
        .options(
            selectinload(Product.plant),
            selectinload(Product.modified_by_user),
        )
        .where(Product.code == code, Product.deleted_at.is_(None))
        .execution_options(populate_existing=True)
    )
    return result.scalar_one_or_none()


async def _ensure_plant(db: AsyncSession, code: int | None, field: str) -> None:
    if code is None:
        return
    obj = await db.scalar(
        select(Object).where(
            Object.code == code,
            Object.deleted_at.is_(None),
            Object.type == "plant",
        )
    )
    if obj is None:
        raise ValueError(f"{field}: завод {code} не найден")


def _touch(product: Product, user: User) -> None:
    product.last_modified_by = user.id
    product.last_modified_at = datetime.now(UTC)


def _audit(
    db: AsyncSession,
    *,
    user: User,
    action: str,
    entity_id: str,
    payload: dict[str, Any],
) -> None:
    add_audit_log(
        db,
        entity_type="product",
        entity_id=entity_id,
        action=action,
        user=user,
        payload=payload,
    )


async def get_product_for_edit(db: AsyncSession, code: int) -> ProductDetail:
    product = await _load_product(db, code)
    if product is None:
        raise APIError(404, "NOT_FOUND", "Продукт не найден")
    return to_product_detail(product)


async def update_product(
    db: AsyncSession, code: int, body: ProductUpdate, user: User
) -> ProductDetail:
    product = await _load_product(db, code)
    if product is None:
        raise APIError(404, "NOT_FOUND", "Продукт не найден")
    try:
        gtin = _normalize_gtin(body.gtin)
        await _ensure_plant(db, body.plant_id, "plant_id")
        await _ensure_plant(db, body.second_plant_id, "second_plant_id")
        await _ensure_plant(db, body.third_plant_id, "third_plant_id")
    except ValueError as exc:
        raise APIError(400, "VALIDATION_ERROR", str(exc)) from exc

    before = {
        "name": product.name,
        "category": product.category.strip(),
        "is_active": product.is_active,
        "gtin": product.gtin,
    }
    product.name = body.name.strip()
    product.description = body.description
    product.category = body.category
    product.is_active = body.is_active
    product.weight_kg = body.weight_kg
    product.monthly_consumption = body.monthly_consumption
    product.gtin = gtin
    product.mark_control = body.mark_control
    product.plant_id = body.plant_id
    product.second_plant_id = body.second_plant_id
    product.third_plant_id = body.third_plant_id
    product.parent_code = body.parent_code
    product.children_code = body.children_code
    _touch(product, user)
    _audit(
        db,
        user=user,
        action="update",
        entity_id=str(code),
        payload={"before": before},
    )
    await db.commit()
    db.expire_all()
    updated = await _load_product(db, code)
    assert updated is not None
    return to_product_detail(updated)


async def delete_product(db: AsyncSession, code: int, user: User) -> None:
    product = await _load_product(db, code)
    if product is None:
        raise APIError(404, "NOT_FOUND", "Продукт не найден")

    has_items = await db.scalar(
        select(exists().where(RequestItem.product_code == code))
    )
    has_normatives = await db.scalar(
        select(
            exists().where(
                Normative.product_code == code,
                Normative.deleted_at.is_(None),
            )
        )
    )
    has_balances = await db.scalar(
        select(exists().where(AvailableBalance.product_code == code))
    )
    if has_items or has_normatives or has_balances:
        raise APIError(
            409,
            "HAS_RELATIONS",
            "Нельзя удалить продукт: есть связанные нормативы, "
            "позиции запросов или остатки",
        )

    product.deleted_at = datetime.now(UTC)
    _touch(product, user)
    _audit(
        db,
        user=user,
        action="delete",
        entity_id=str(code),
        payload={"name": product.name},
    )
    await db.commit()


async def _upsert_row(
    db: AsyncSession,
    user: User,
    values: dict[str, Any],
) -> str:
    code = values["code"]
    gtin = values.get("gtin")
    product = await db.scalar(
        select(Product).where(Product.code == code, Product.deleted_at.is_(None))
    )
    if product is None:
        plant_id = values.get("plant_id")
        weight = values.get("weight_kg")
        if plant_id is None:
            raise ValueError("Для нового продукта укажите plant_id")
        if weight is None or weight <= 0:
            raise ValueError("Для нового продукта укажите weight_kg > 0")
        await _ensure_plant(db, plant_id, "plant_id")
        if not values.get("name"):
            raise ValueError("Укажите name")
        category = values.get("category")
        if category not in {"A", "B", "C"}:
            raise ValueError("category должна быть A, B или C")
        product = Product(
            code=code,
            name=values["name"],
            category=category,
            plant_id=plant_id,
            weight_kg=weight,
            is_active=values["is_active"],
            parent_code=values.get("parent_code"),
            children_code=values.get("children_code"),
            gtin=gtin,
            mark_control=values["mark_control"],
        )
        _touch(product, user)
        db.add(product)
        _audit(
            db,
            user=user,
            action="create",
            entity_id=str(code),
            payload={"source": "excel"},
        )
        return "created"

    if values.get("name"):
        product.name = values["name"]
    if values.get("category") in {"A", "B", "C"}:
        product.category = values["category"]
    product.is_active = values["is_active"]
    product.mark_control = values["mark_control"]
    if "gtin" in values:
        product.gtin = gtin
    if values.get("parent_code") is not None or values.get("clear_parent"):
        product.parent_code = values.get("parent_code")
    if values.get("children_code") is not None or values.get("clear_children"):
        product.children_code = values.get("children_code")
    if values.get("plant_id") is not None:
        await _ensure_plant(db, values["plant_id"], "plant_id")
        product.plant_id = values["plant_id"]
    if values.get("weight_kg") is not None:
        if values["weight_kg"] <= 0:
            raise ValueError("weight_kg должен быть больше 0")
        product.weight_kg = values["weight_kg"]
    _touch(product, user)
    _audit(
        db,
        user=user,
        action="update",
        entity_id=str(code),
        payload={"source": "excel"},
    )
    return "updated"


async def upload_products(
    db: AsyncSession, user: User, content: bytes, filename: str
) -> ProductUploadResult:
    lowered = filename.lower()
    if not lowered.endswith((".xlsx", ".xls")):
        raise APIError(400, "INVALID_FILE", "Загрузите файл .xlsx или .xls")
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise APIError(
            400,
            "INVALID_FILE",
            "Не удалось прочитать Excel. Используйте формат .xlsx",
        ) from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise APIError(400, "INVALID_FILE", "Файл пуст")
    headers = _cell_map(rows[0])
    if "code" not in headers:
        raise APIError(400, "INVALID_FILE", "В шаблоне нет колонки code")

    created = 0
    updated = 0
    error_details: list[ProductUploadError] = []

    for index, row in enumerate(rows[1:], start=2):
        if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        try:
            code = _parse_int(row[headers["code"]], "code")
            if code is None:
                raise ValueError("Укажите code")
            name_idx = headers.get("name")
            category_raw = (
                str(row[headers["category"]]).strip().upper()
                if "category" in headers and row[headers["category"]] is not None
                else ""
            )
            values: dict[str, Any] = {
                "code": code,
                "name": (
                    str(row[name_idx]).strip()
                    if name_idx is not None and row[name_idx] is not None
                    else ""
                ),
                "is_active": _parse_active(
                    row[headers["status"]] if "status" in headers else None
                ),
                "mark_control": _parse_bool_mark(
                    row[headers["mark_control"]] if "mark_control" in headers else None
                ),
                "gtin": _normalize_gtin(
                    row[headers["gtin"]] if "gtin" in headers else None
                ),
                "category": category_raw or None,
                "parent_code": _parse_int(
                    row[headers["parent_code"]] if "parent_code" in headers else None,
                    "parent_code",
                ),
                "children_code": _parse_int(
                    (
                        row[headers["children_code"]]
                        if "children_code" in headers
                        else None
                    ),
                    "children_code",
                ),
                "plant_id": _parse_int(
                    row[headers["plant_id"]] if "plant_id" in headers else None,
                    "plant_id",
                ),
                "weight_kg": _parse_decimal(
                    row[headers["weight_kg"]] if "weight_kg" in headers else None,
                    "weight_kg",
                ),
            }
            if "parent_code" in headers:
                raw_parent = row[headers["parent_code"]]
                values["clear_parent"] = (
                    raw_parent is None or str(raw_parent).strip() == ""
                )
            if "children_code" in headers:
                raw_child = row[headers["children_code"]]
                values["clear_children"] = (
                    raw_child is None or str(raw_child).strip() == ""
                )
            async with db.begin_nested():
                result = await _upsert_row(db, user, values)
            if result == "created":
                created += 1
            else:
                updated += 1
        except (ValueError, APIError) as exc:
            message = exc.message if isinstance(exc, APIError) else str(exc)
            error_details.append(ProductUploadError(row=index, message=message))

    _audit(
        db,
        user=user,
        action="upload",
        entity_id=filename,
        payload={
            "created": created,
            "updated": updated,
            "errors": len(error_details),
            "filename": filename,
        },
    )
    await db.commit()
    loaded = created + updated
    return ProductUploadResult(
        created=created,
        updated=updated,
        errors=len(error_details),
        message=f"Загружено {loaded}, ошибок {len(error_details)}",
        error_details=error_details,
    )
