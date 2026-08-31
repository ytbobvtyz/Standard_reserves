import logging
from collections import defaultdict
from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Literal
from uuid import UUID

from openpyxl import Workbook, load_workbook
from sqlalchemy import case, delete, exists, or_, select, tuple_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.exceptions import APIError
from app.models.available_balance import AvailableBalance
from app.models.normative import Normative
from app.models.object import Object
from app.models.product import Product
from app.models.production_request import ProductionRequest, ProductionRequestItem
from app.models.sync_metadata import SyncMetadata
from app.models.user import User
from app.schemas.logistics import (
    BalanceSyncInfo,
    BalanceSyncUser,
    BalanceUploadError,
    BalanceUploadResult,
    DashboardResponse,
    DashboardSummary,
    DeficitItem,
    FilterMode,
    GenerateOrdersData,
    OrderItem,
    PlantOrder,
    Unit,
    WarehouseDeficit,
)
from app.services.coefficients import calculate_requirement

logger = logging.getLogger(__name__)

ESTIMATED_DELIVERY_DAYS = 5
KG_IN_TON = Decimal("1000")
LONG_DISTANCE_MESSAGE = (
    "Ввиду удалённого расположения склада, пополнение возможно по железной дороге "
    "— срок доставки около 1 месяца от даты готовности продукции на производственной "
    "площадке, в связи с чем нормативы увеличены"
)


@dataclass
class DeficitRow:
    warehouse_code: int
    warehouse_name: str
    product_code: int
    product_name: str
    category: str
    plant_id: int
    second_plant_id: int | None
    third_plant_id: int | None
    plant_name: str
    weight_kg: Decimal
    normative_quantity: Decimal
    requirement: Decimal
    available: Decimal
    plan: Decimal
    deficit: Decimal
    unit: str
    client_name: str
    expiry_date: date | None
    status: Literal["warning", "ok"]
    stock_unit: str
    long_distance: bool
    is_active: bool = True
    parent_code: int | None = None
    children_code: int | None = None
    group_key: str = ""
    group_index: int = 0
    is_group_main: bool = True
    hide_group_metrics: bool = False


def _quantize(value: Decimal, unit: Unit) -> Decimal:
    step = Decimal("0.0001") if unit == "т" else Decimal("0.01")
    return value.quantize(step, rounding=ROUND_HALF_UP)


def _normalize_stock_unit(unit: str | None) -> str:
    text = (unit or "шт").strip()
    if text.upper() == "ШТ":
        return "ШТ"
    if text.upper() == "КГ":
        return "КГ"
    if text == "т":
        return "т"
    return text or "шт"


def to_pieces(quantity: Decimal, unit: str, weight_kg: Decimal) -> Decimal:
    normalized = _normalize_stock_unit(unit)
    if normalized in {"шт", "ШТ"}:
        return quantity
    if weight_kg <= 0:
        return Decimal("0")
    if normalized == "КГ":
        return quantity / weight_kg
    return quantity * KG_IN_TON / weight_kg


def from_pieces(quantity: Decimal, unit: Unit, weight_kg: Decimal) -> Decimal:
    if unit == "шт":
        return quantity
    return quantity * weight_kg / KG_IN_TON


def resolve_plant_id(product: Product) -> int:
    for code in (product.plant_id, product.second_plant_id, product.third_plant_id):
        if code:
            return code
    return product.plant_id


def _item_key(warehouse_code: int, product_code: int) -> tuple[int, int]:
    return (warehouse_code, product_code)


def _is_group_main(product: Product) -> bool:
    return bool(product.is_active) and product.children_code is None


def _pick_group_main(members: list[Product]) -> Product:
    mains = [item for item in members if _is_group_main(item)]
    if mains:
        return min(mains, key=lambda item: item.code)
    tips = [item for item in members if item.children_code is None]
    if tips:
        active = [item for item in tips if item.is_active]
        return min(active or tips, key=lambda item: item.code)
    return min(members, key=lambda item: item.code)


def _member_to_main(products: dict[int, Product]) -> dict[int, int]:
    if not products:
        return {}
    parent: dict[int, int] = {code: code for code in products}

    def find(code: int) -> int:
        while parent[code] != code:
            parent[code] = parent[parent[code]]
            code = parent[code]
        return code

    def union(left: int, right: int) -> None:
        if left not in parent or right not in parent:
            return
        root_left, root_right = find(left), find(right)
        if root_left != root_right:
            parent[root_right] = root_left

    for product in products.values():
        if product.parent_code:
            union(product.code, product.parent_code)
        if product.children_code:
            union(product.code, product.children_code)

    families: dict[int, list[Product]] = defaultdict(list)
    for product in products.values():
        families[find(product.code)].append(product)

    mapping: dict[int, int] = {}
    for members in families.values():
        main_code = _pick_group_main(members).code
        for product in members:
            mapping[product.code] = main_code
    return mapping


async def _load_product_map(
    db: AsyncSession, seed_codes: set[int]
) -> dict[int, Product]:
    products: dict[int, Product] = {}
    pending = set(seed_codes)
    while pending:
        result = await db.execute(
            select(Product)
            .options(selectinload(Product.plant))
            .where(
                Product.deleted_at.is_(None),
                or_(
                    Product.code.in_(pending),
                    Product.parent_code.in_(pending),
                    Product.children_code.in_(pending),
                ),
            )
        )
        pending = set()
        for product in result.scalars().unique().all():
            if product.code in products:
                continue
            products[product.code] = product
            for related in (product.parent_code, product.children_code):
                if related and related not in products:
                    pending.add(related)
    return products


async def _expand_family_codes(db: AsyncSession, codes: list[int]) -> list[int]:
    products = await _load_product_map(db, set(codes))
    if not products:
        return list(dict.fromkeys(codes))
    mapping = _member_to_main(products)
    mains = {mapping.get(code, code) for code in codes}
    family = {code for code, main in mapping.items() if main in mains}
    family.update(codes)
    return sorted(family)


async def _expand_related_members(
    db: AsyncSession,
    aggregated: dict[tuple[int, int], dict],
    balances: dict[tuple[int, int], AvailableBalance],
    products: dict[int, Product],
) -> None:
    if not aggregated or not products:
        return
    member_to_main = _member_to_main(products)
    warehouses: dict[int, Object] = {}
    present_by_warehouse: dict[int, set[int]] = defaultdict(set)
    for (warehouse_code, product_code), bucket in aggregated.items():
        warehouses[warehouse_code] = bucket["warehouse"]
        present_by_warehouse[warehouse_code].add(product_code)

    extra_keys: set[tuple[int, int]] = set()
    for warehouse_code, codes in present_by_warehouse.items():
        mains = {member_to_main.get(code, code) for code in codes}
        family_codes = {
            code for code, main in member_to_main.items() if main in mains
        } | codes
        for product_code in family_codes:
            if (warehouse_code, product_code) not in aggregated:
                extra_keys.add((warehouse_code, product_code))

    if extra_keys:
        balances.update(await _load_balances(db, keys=extra_keys))

    for warehouse_code, product_code in extra_keys:
        product = products.get(product_code)
        if product is None:
            continue
        has_balance = (warehouse_code, product_code) in balances
        if not _is_group_main(product) and not has_balance:
            continue
        aggregated[(warehouse_code, product_code)] = {
            "warehouse": warehouses[warehouse_code],
            "product": product,
            "normative_pcs": Decimal("0"),
            "clients": [],
            "expiry_date": None,
        }


def _convert_row(
    *,
    warehouse: Object,
    product: Product,
    plant_name: str,
    normative_pcs: Decimal,
    available_pcs: Decimal,
    plan_pcs: Decimal,
    unit: Unit,
    client_name: str,
    expiry_date: date | None,
    stock_unit: str,
    long_distance: bool | None = None,
    deficit_plan_pcs: Decimal | None = None,
    group_key: str | None = None,
    group_index: int = 0,
    is_group_main: bool = True,
    hide_group_metrics: bool = False,
) -> DeficitRow:
    is_remote = bool(
        warehouse.long_distance if long_distance is None else long_distance
    )
    if hide_group_metrics:
        requirement_pcs = Decimal("0")
        deficit_pcs = Decimal("0")
        status: Literal["warning", "ok"] = "ok"
        display_normative_pcs = Decimal("0")
    else:
        display_normative_pcs = normative_pcs
        requirement_pcs = calculate_requirement(
            display_normative_pcs, product.category, is_remote
        )
        plan_for_deficit = plan_pcs if deficit_plan_pcs is None else deficit_plan_pcs
        deficit_pcs = requirement_pcs - plan_for_deficit
        status = "warning" if deficit_pcs > 0 else "ok"
    return DeficitRow(
        warehouse_code=warehouse.code,
        warehouse_name=warehouse.name,
        product_code=product.code,
        product_name=product.name,
        category=product.category.strip(),
        plant_id=resolve_plant_id(product),
        second_plant_id=product.second_plant_id,
        third_plant_id=product.third_plant_id,
        plant_name=plant_name,
        weight_kg=product.weight_kg,
        normative_quantity=_quantize(
            from_pieces(display_normative_pcs, unit, product.weight_kg), unit
        ),
        requirement=_quantize(
            from_pieces(requirement_pcs, unit, product.weight_kg), unit
        ),
        available=_quantize(from_pieces(available_pcs, unit, product.weight_kg), unit),
        plan=_quantize(from_pieces(plan_pcs, unit, product.weight_kg), unit),
        deficit=_quantize(from_pieces(deficit_pcs, unit, product.weight_kg), unit),
        unit=unit,
        client_name=client_name,
        expiry_date=expiry_date,
        status=status,
        stock_unit=stock_unit,
        long_distance=is_remote,
        is_active=bool(product.is_active),
        parent_code=product.parent_code,
        children_code=product.children_code,
        group_key=group_key or str(product.code),
        group_index=group_index,
        is_group_main=is_group_main,
        hide_group_metrics=hide_group_metrics,
    )


async def _load_plants(db: AsyncSession, plant_ids: set[int]) -> dict[int, Object]:
    if not plant_ids:
        return {}
    result = await db.execute(select(Object).where(Object.code.in_(plant_ids)))
    return {item.code: item for item in result.scalars().all()}


async def _load_balances(
    db: AsyncSession,
    *,
    keys: set[tuple[int, int]] | None = None,
    warehouse_code: int | None = None,
    warehouse_codes: list[int] | None = None,
    product_codes: list[int] | None = None,
) -> dict[tuple[int, int], AvailableBalance]:
    conditions = []
    if keys is not None:
        if not keys:
            return {}
        conditions.append(
            tuple_(
                AvailableBalance.warehouse_code,
                AvailableBalance.product_code,
            ).in_(list(keys))
        )
    else:
        if warehouse_code is not None:
            conditions.append(AvailableBalance.warehouse_code == warehouse_code)
        elif warehouse_codes:
            conditions.append(AvailableBalance.warehouse_code.in_(warehouse_codes))
        if product_codes:
            conditions.append(AvailableBalance.product_code.in_(product_codes))
    stmt = select(AvailableBalance)
    if conditions:
        stmt = stmt.where(*conditions)
    return {
        (item.warehouse_code, item.product_code): item
        for item in (await db.execute(stmt)).scalars().all()
    }


async def collect_deficit_rows(
    db: AsyncSession,
    *,
    warehouse_code: int | None = None,
    warehouse_codes: list[int] | None = None,
    filter_mode: FilterMode = "with_normatives",
    unit: Unit = "шт",
    product_codes: list[int] | None = None,
) -> list[DeficitRow]:
    if product_codes:
        product_codes = await _expand_family_codes(db, product_codes)
    today = date.today()
    conditions = [
        Normative.deleted_at.is_(None),
        Normative.expiry_date >= today,
        or_(
            Normative.production_request_item_id.is_(None),
            exists(
                select(1)
                .select_from(ProductionRequestItem)
                .join(
                    ProductionRequest,
                    ProductionRequest.id == ProductionRequestItem.production_request_id,
                )
                .where(
                    ProductionRequestItem.id == Normative.production_request_item_id,
                    ProductionRequest.deleted_at.is_(None),
                    ProductionRequest.status == "active",
                    ProductionRequest.valid_from <= today,
                )
            ),
        ),
        Product.deleted_at.is_(None),
        Object.deleted_at.is_(None),
    ]
    if warehouse_code is not None:
        conditions.append(Normative.warehouse_code == warehouse_code)
    elif warehouse_codes:
        conditions.append(Normative.warehouse_code.in_(warehouse_codes))
    if product_codes:
        conditions.append(Normative.product_code.in_(product_codes))

    result = await db.execute(
        select(Normative)
        .options(
            selectinload(Normative.product).selectinload(Product.plant),
            selectinload(Normative.warehouse),
        )
        .join(Product, Product.code == Normative.product_code)
        .join(Object, Object.code == Normative.warehouse_code)
        .where(*conditions)
        .order_by(Normative.warehouse_code, Normative.product_code)
    )
    normatives = list(result.scalars().unique().all())

    aggregated: dict[tuple[int, int], dict] = {}
    for normative in normatives:
        product = normative.product
        key = _item_key(normative.warehouse_code, normative.product_code)
        bucket = aggregated.setdefault(
            key,
            {
                "warehouse": normative.warehouse,
                "product": product,
                "normative_pcs": Decimal("0"),
                "clients": [],
                "expiry_date": normative.expiry_date,
            },
        )
        bucket["normative_pcs"] += to_pieces(
            normative.quantity, normative.unit, product.weight_kg
        )
        if normative.client_name and normative.client_name not in bucket["clients"]:
            bucket["clients"].append(normative.client_name)
        if (
            bucket["expiry_date"] is None
            or normative.expiry_date < bucket["expiry_date"]
        ):
            bucket["expiry_date"] = normative.expiry_date

    if filter_mode == "all":
        balances = await _load_balances(
            db,
            warehouse_code=warehouse_code,
            warehouse_codes=warehouse_codes,
            product_codes=product_codes,
        )
    else:
        balances = await _load_balances(db, keys=set(aggregated))

    covered_keys = set(aggregated)
    if filter_mode == "all":
        extra_keys = [key for key in balances if key not in covered_keys]
        extra_product_codes = {product_code for _, product_code in extra_keys}
        extra_warehouse_codes = {wh for wh, _ in extra_keys}
        products_by_code: dict[int, Product] = {}
        warehouses_by_code: dict[int, Object] = {}
        if extra_product_codes:
            products_by_code = {
                item.code: item
                for item in (
                    await db.execute(
                        select(Product)
                        .options(selectinload(Product.plant))
                        .where(
                            Product.code.in_(extra_product_codes),
                            Product.deleted_at.is_(None),
                        )
                    )
                )
                .scalars()
                .all()
            }
        if extra_warehouse_codes:
            warehouses_by_code = {
                item.code: item
                for item in (
                    await db.execute(
                        select(Object).where(
                            Object.code.in_(extra_warehouse_codes),
                            Object.deleted_at.is_(None),
                        )
                    )
                )
                .scalars()
                .all()
            }
        for warehouse_code_key, product_code in extra_keys:
            product = products_by_code.get(product_code)
            warehouse = warehouses_by_code.get(warehouse_code_key)
            if product is None or warehouse is None:
                continue
            aggregated[(warehouse_code_key, product_code)] = {
                "warehouse": warehouse,
                "product": product,
                "normative_pcs": Decimal("0"),
                "clients": [],
                "expiry_date": None,
            }

    products_by_code = {
        bucket["product"].code: bucket["product"] for bucket in aggregated.values()
    }
    if any(
        product.parent_code or product.children_code
        for product in products_by_code.values()
    ):
        products_by_code = await _load_product_map(db, set(products_by_code))
        await _expand_related_members(db, aggregated, balances, products_by_code)

    plant_ids = {resolve_plant_id(bucket["product"]) for bucket in aggregated.values()}
    plants = await _load_plants(db, plant_ids)
    warehouse_codes_used = {bucket["warehouse"].code for bucket in aggregated.values()}
    long_distance_by_code: dict[int, bool] = {}
    if warehouse_codes_used:
        flags = await db.execute(
            select(Object.code, Object.long_distance).where(
                Object.code.in_(warehouse_codes_used)
            )
        )
        long_distance_by_code = {code: bool(flag) for code, flag in flags.all()}

    member_to_main = _member_to_main(products_by_code)
    groups: dict[tuple[int, int], list[tuple[int, int]]] = defaultdict(list)
    for warehouse_code_key, product_code in aggregated:
        main_code = member_to_main.get(product_code, product_code)
        groups[(warehouse_code_key, main_code)].append(
            (warehouse_code_key, product_code)
        )

    rows: list[DeficitRow] = []
    group_index_by_warehouse: dict[int, int] = defaultdict(int)
    for warehouse_code_key, main_code in sorted(groups):
        member_keys = sorted(
            groups[(warehouse_code_key, main_code)],
            key=lambda item: (0 if item[1] == main_code else 1, item[1]),
        )
        group_normative_pcs = sum(
            (aggregated[key]["normative_pcs"] for key in member_keys),
            Decimal("0"),
        )
        group_plan_pcs = Decimal("0")
        group_clients: list[str] = []
        group_expiry: date | None = None
        for key in member_keys:
            bucket = aggregated[key]
            product = bucket["product"]
            balance = balances.get(key)
            if balance is not None:
                group_plan_pcs += to_pieces(
                    balance.plan, balance.unit, product.weight_kg
                )
            for client_name in bucket["clients"]:
                if client_name and client_name not in group_clients:
                    group_clients.append(client_name)
            expiry = bucket["expiry_date"]
            if expiry is not None and (group_expiry is None or expiry < group_expiry):
                group_expiry = expiry

        main_bucket = aggregated.get(
            (warehouse_code_key, main_code), aggregated[member_keys[0]]
        )
        main_product: Product = main_bucket["product"]
        main_warehouse: Object = main_bucket["warehouse"]
        is_remote = long_distance_by_code.get(
            main_warehouse.code, bool(main_warehouse.long_distance)
        )
        group_requirement_pcs = calculate_requirement(
            group_normative_pcs, main_product.category, is_remote
        )
        group_deficit_pcs = group_requirement_pcs - group_plan_pcs
        if filter_mode == "with_normatives" and group_normative_pcs <= 0:
            continue
        if filter_mode == "deficit_only" and group_deficit_pcs <= 0:
            continue

        group_index = group_index_by_warehouse[warehouse_code_key]
        group_index_by_warehouse[warehouse_code_key] += 1
        grouped = len(member_keys) > 1

        for key in member_keys:
            bucket = aggregated[key]
            product = bucket["product"]
            warehouse: Object = bucket["warehouse"]
            balance = balances.get(key)
            available_pcs = (
                to_pieces(balance.available, balance.unit, product.weight_kg)
                if balance is not None
                else Decimal("0")
            )
            plan_pcs = (
                to_pieces(balance.plan, balance.unit, product.weight_kg)
                if balance is not None
                else Decimal("0")
            )
            plant = plants.get(resolve_plant_id(product))
            is_main = product.code == main_code
            rows.append(
                _convert_row(
                    warehouse=warehouse,
                    product=product,
                    plant_name=(
                        plant.name if plant else f"Завод {resolve_plant_id(product)}"
                    ),
                    normative_pcs=group_normative_pcs if is_main else Decimal("0"),
                    available_pcs=available_pcs,
                    plan_pcs=plan_pcs,
                    unit=unit,
                    client_name=", ".join(
                        group_clients if is_main else bucket["clients"]
                    ),
                    expiry_date=group_expiry if is_main else bucket["expiry_date"],
                    stock_unit=(
                        _normalize_stock_unit(balance.unit)
                        if balance is not None
                        else "ШТ"
                    ),
                    long_distance=long_distance_by_code.get(
                        warehouse.code, bool(warehouse.long_distance)
                    ),
                    deficit_plan_pcs=group_plan_pcs if is_main else Decimal("0"),
                    group_key=str(main_code),
                    group_index=group_index,
                    is_group_main=is_main,
                    hide_group_metrics=grouped and not is_main,
                )
            )
    return rows


def _to_deficit_item(row: DeficitRow) -> DeficitItem:
    return DeficitItem(
        product_code=row.product_code,
        product_name=row.product_name,
        category=row.category,
        normative_quantity=row.normative_quantity,
        requirement=row.requirement,
        available=row.available,
        plan=row.plan,
        unit=row.unit,
        deficit=row.deficit,
        client_name=row.client_name,
        expiry_date=row.expiry_date,
        status=row.status,
        stock_unit=row.stock_unit,
        weight_kg=row.weight_kg,
        is_active=row.is_active,
        parent_code=row.parent_code,
        children_code=row.children_code,
        group_key=row.group_key or str(row.product_code),
        group_index=row.group_index,
        is_group_main=row.is_group_main,
        hide_group_metrics=row.hide_group_metrics,
    )


def build_dashboard(rows: list[DeficitRow]) -> DashboardResponse:
    grouped: dict[int, list[DeficitRow]] = defaultdict(list)
    names: dict[int, str] = {}
    remote: dict[int, bool] = {}
    for row in rows:
        grouped[row.warehouse_code].append(row)
        names[row.warehouse_code] = row.warehouse_name
        remote[row.warehouse_code] = row.long_distance

    warehouses: list[WarehouseDeficit] = []
    for warehouse_code in sorted(grouped):
        items = grouped[warehouse_code]
        positive = [item for item in items if item.deficit > 0]
        total = sum((item.deficit for item in positive), Decimal("0"))
        is_remote = remote.get(warehouse_code, False)
        warehouses.append(
            WarehouseDeficit(
                warehouse_code=warehouse_code,
                warehouse_name=names[warehouse_code],
                long_distance=is_remote,
                long_distance_message=LONG_DISTANCE_MESSAGE if is_remote else None,
                deficit_items=[_to_deficit_item(item) for item in items],
                total_deficit=total,
                deficit_count=len(positive),
            )
        )

    positive_rows = [row for row in rows if row.deficit > 0]
    return DashboardResponse(
        data=warehouses,
        summary=DashboardSummary(
            total_deficit=sum((row.deficit for row in positive_rows), Decimal("0")),
            deficit_warehouses=len({row.warehouse_code for row in positive_rows}),
            deficit_products=len({row.product_code for row in positive_rows}),
        ),
    )


async def get_dashboard(
    db: AsyncSession,
    *,
    warehouse_code: int | None = None,
    filter_mode: FilterMode = "with_normatives",
    unit: Unit = "шт",
) -> DashboardResponse:
    rows = await collect_deficit_rows(
        db,
        warehouse_code=warehouse_code,
        filter_mode=filter_mode,
        unit=unit,
    )
    return build_dashboard(rows)


def _orders_payload(orders: list[PlantOrder]) -> GenerateOrdersData:
    product_codes_count = {item.product_code for row in orders for item in row.items}
    total_quantity = sum(
        (item.deficit for row in orders for item in row.items),
        Decimal("0"),
    )
    return GenerateOrdersData(
        orders=orders,
        total_orders=len(orders),
        total_products=len(product_codes_count),
        total_quantity=total_quantity,
    )


async def _load_warehouses(
    db: AsyncSession, warehouse_codes: list[int]
) -> dict[int, Object]:
    if not warehouse_codes:
        return {}
    result = await db.execute(
        select(Object).where(
            Object.code.in_(warehouse_codes),
            Object.deleted_at.is_(None),
            Object.type == "warehouse",
        )
    )
    return {item.code: item for item in result.scalars().all()}


async def generate_orders(
    db: AsyncSession,
    *,
    warehouse_code: int,
    product_codes: list[int] | None = None,
) -> GenerateOrdersData:
    return await generate_orders_for_warehouses(
        db,
        warehouse_codes=[warehouse_code],
        product_codes=product_codes,
    )


async def generate_orders_for_warehouses(
    db: AsyncSession,
    *,
    warehouse_codes: list[int],
    product_codes: list[int] | None = None,
) -> GenerateOrdersData:
    if not warehouse_codes:
        raise APIError(400, "VALIDATION_ERROR", "Укажите хотя бы один склад")

    unique_codes = list(dict.fromkeys(warehouse_codes))
    warehouses = await _load_warehouses(db, unique_codes)
    for code in unique_codes:
        if code not in warehouses:
            raise APIError(404, "NOT_FOUND", "Склад не найден")

    rows = await collect_deficit_rows(
        db,
        warehouse_codes=unique_codes,
        filter_mode="deficit_only",
        unit="шт",
        product_codes=product_codes,
    )
    grouped: dict[tuple[int, int], list[DeficitRow]] = defaultdict(list)
    for row in rows:
        if row.deficit <= 0:
            continue
        grouped[(row.warehouse_code, row.plant_id)].append(row)

    warehouse_order = {code: index for index, code in enumerate(unique_codes)}
    plant_ids = {plant_id for _, plant_id in grouped}
    plants = await _load_plants(db, plant_ids)
    orders: list[PlantOrder] = []
    for warehouse_code, plant_code in sorted(
        grouped,
        key=lambda item: (warehouse_order[item[0]], item[1]),
    ):
        warehouse = warehouses[warehouse_code]
        plant = plants.get(plant_code)
        items = grouped[(warehouse_code, plant_code)]
        orders.append(
            PlantOrder(
                plant_code=plant_code,
                plant_name=plant.name if plant else items[0].plant_name,
                warehouse_code=warehouse.code,
                warehouse_name=warehouse.name,
                items=[
                    OrderItem(
                        product_code=item.product_code,
                        product_name=item.product_name,
                        deficit=item.deficit,
                        unit=item.unit,
                        weight_kg=item.weight_kg,
                    )
                    for item in items
                ],
                estimated_delivery_days=ESTIMATED_DELIVERY_DAYS,
            )
        )

    return _orders_payload(orders)


async def export_excel(
    db: AsyncSession,
    *,
    warehouse_code: int | None = None,
    product_codes: list[int] | None = None,
    unit: Unit = "шт",
) -> bytes:
    rows = await collect_deficit_rows(
        db,
        warehouse_code=warehouse_code,
        filter_mode="deficit_only",
        unit=unit,
        product_codes=product_codes,
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Дефицит"
    sheet.append(
        [
            "Склад",
            "Артикул",
            "Название",
            "Норматив",
            "Потребность",
            "Доступно",
            "Запланировано",
            "Дефицит",
            "Ед",
            "Клиент",
        ]
    )
    for row in rows:
        if row.hide_group_metrics:
            continue
        sheet.append(
            [
                row.warehouse_name,
                row.product_code,
                row.product_name,
                float(row.normative_quantity),
                float(row.requirement),
                float(row.available),
                float(row.plan),
                float(row.deficit),
                row.unit,
                row.client_name,
            ]
        )
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


EXCEL_COL_PRODUCT = 1
EXCEL_COL_PLANT = 3
EXCEL_COL_WAREHOUSE = 5
EXCEL_COL_UNIT = 17
EXCEL_COL_AVAILABLE = 19
EXCEL_COL_PLAN = 20
HEADER_MARKERS = {
    "product_code",
    "артикул",
    "код",
    "material",
    "мат",
    "код материала",
}


def _cell(row: tuple[Any, ...], column: int) -> Any:
    index = column - 1
    if index < 0 or index >= len(row):
        return None
    return row[index]


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    text = str(value).strip()
    return text == "" or text.lower() in {"none", "nan"}


def _looks_like_header(value: Any) -> bool:
    if _is_blank(value):
        return False
    text = str(value).strip().lower()
    if text in HEADER_MARKERS:
        return True
    try:
        float(text.replace(",", "."))
        return False
    except ValueError:
        return any(char.isalpha() for char in text)


def _parse_int_cell(value: Any, field: str) -> int:
    if _is_blank(value):
        raise ValueError(f"Укажите {field}")
    if isinstance(value, bool):
        raise ValueError(f"{field} должен быть числом")
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        as_int = int(value)
        if float(as_int) == value:
            return as_int
        raise ValueError(f"{field} должен быть целым числом")
    if isinstance(value, Decimal):
        as_int = int(value)
        if Decimal(as_int) == value:
            return as_int
        raise ValueError(f"{field} должен быть целым числом")
    text = str(value).strip().replace(",", ".")
    if text.endswith(".0"):
        text = text[:-2]
    try:
        parsed = float(text)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field} должен быть числом") from exc
    as_int = int(parsed)
    if parsed != float(as_int):
        raise ValueError(f"{field} должен быть целым числом")
    return as_int


def _parse_warehouse_code(value: Any) -> str:
    if _is_blank(value):
        raise ValueError("Укажите склад (erp_warehouse_code)")
    if isinstance(value, bool):
        raise ValueError("erp_warehouse_code должен содержать до 4 символов")
    if isinstance(value, int):
        text = str(value)
    elif isinstance(value, float):
        as_int = int(value)
        if float(as_int) != value:
            raise ValueError("erp_warehouse_code должен содержать до 4 символов")
        text = str(as_int)
    else:
        text = str(value).strip().upper()
        if text.endswith(".0") and text[:-2].replace(".", "", 1).isdigit():
            text = text[:-2]
        try:
            as_number = int(float(text.replace(",", ".")))
            if text.replace(",", ".") in {str(as_number), f"{as_number}.0"}:
                text = str(as_number)
        except ValueError:
            pass
    if len(text) > 4:
        raise ValueError("erp_warehouse_code должен содержать до 4 символов")
    return text


def _parse_balance_unit(value: Any) -> str:
    if _is_blank(value):
        return "ШТ"
    text = str(value).strip().upper().replace(" ", "")
    if text in {"ШТ", "ШТ."}:
        return "ШТ"
    if text in {"КГ", "КГ."}:
        return "КГ"
    raise ValueError("Единица измерения должна быть ШТ или КГ")


def _parse_quantity(value: Any, field: str) -> Decimal:
    if _is_blank(value):
        raise ValueError(f"Укажите {field}")
    if isinstance(value, bool):
        raise ValueError(f"{field} должен быть числом")
    if isinstance(value, (int, float, Decimal)):
        amount = Decimal(str(value))
    else:
        text = str(value).strip().replace(" ", "").replace(",", ".")
        try:
            amount = Decimal(text)
        except (InvalidOperation, TypeError, ValueError) as exc:
            raise ValueError(f"{field} должен быть числом") from exc
    return amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _row_empty(row: tuple[Any, ...]) -> bool:
    tracked = (
        EXCEL_COL_PRODUCT,
        EXCEL_COL_PLANT,
        EXCEL_COL_WAREHOUSE,
        EXCEL_COL_AVAILABLE,
        EXCEL_COL_PLAN,
    )
    return all(_is_blank(_cell(row, column)) for column in tracked)


@dataclass
class ParsedBalanceRow:
    warehouse_code: int
    product_code: int
    available: Decimal
    plan: Decimal
    unit: str


async def _find_plant(db: AsyncSession, erp_plant_code: int) -> Object:
    logger.info(
        "Plant lookup: searching objects.erp_plant_code=%s (int)",
        erp_plant_code,
    )
    result = await db.execute(
        select(Object)
        .where(
            Object.erp_plant_code == erp_plant_code,
            Object.deleted_at.is_(None),
        )
        .order_by(
            case((Object.type == "plant", 0), else_=1),
            Object.is_active.desc(),
            Object.code.asc(),
        )
        .limit(1)
    )
    plant = result.scalars().first()
    if plant is None:
        logger.warning(
            "Plant lookup: no object with erp_plant_code=%s "
            "(not searching objects.code)",
            erp_plant_code,
        )
        raise ValueError(f"Завод ERP {erp_plant_code} не найден")
    logger.info(
        "Plant lookup: found objects.code=%s name=%r type=%s "
        "erp_plant_code=%s is_active=%s",
        plant.code,
        plant.name,
        plant.type,
        plant.erp_plant_code,
        plant.is_active,
    )
    return plant


async def _find_warehouse(db: AsyncSession, erp_warehouse_code: str) -> Object:
    warehouse_key = str(erp_warehouse_code)
    logger.info(
        "Warehouse lookup: searching objects.erp_warehouse_code=%r (str)",
        warehouse_key,
    )
    result = await db.execute(
        select(Object)
        .where(
            Object.erp_warehouse_code == warehouse_key,
            Object.type == "warehouse",
            Object.deleted_at.is_(None),
        )
        .order_by(Object.is_active.desc(), Object.code.asc())
        .limit(1)
    )
    warehouse = result.scalars().first()
    if warehouse is None:
        logger.warning(
            "Warehouse lookup: no warehouse with erp_warehouse_code=%r",
            warehouse_key,
        )
        raise ValueError(f"Склад ERP {warehouse_key} не найден")
    logger.info(
        "Warehouse lookup: found objects.code=%s name=%r type=%s erp_warehouse_code=%r",
        warehouse.code,
        warehouse.name,
        warehouse.type,
        warehouse.erp_warehouse_code,
    )
    return warehouse


async def _find_product(db: AsyncSession, product_code: int) -> Product:
    logger.info("Product lookup: searching products.code=%s (int)", product_code)
    product = await db.scalar(
        select(Product).where(
            Product.code == product_code,
            Product.deleted_at.is_(None),
        )
    )
    if product is None:
        logger.warning("Product lookup: no product with code=%s", product_code)
        raise ValueError(f"Продукт {product_code} не найден")
    logger.info(
        "Product lookup: found products.code=%s name=%r",
        product.code,
        product.name,
    )
    return product


async def _touch_sync_metadata(
    db: AsyncSession, *, user_id: UUID, synced_at: datetime
) -> None:
    row = await db.get(SyncMetadata, 1)
    if row is None:
        db.add(
            SyncMetadata(
                id=1,
                last_balances_sync_at=synced_at,
                last_balances_sync_by=user_id,
                updated_at=synced_at,
            )
        )
        return
    row.last_balances_sync_at = synced_at
    row.last_balances_sync_by = user_id
    row.updated_at = synced_at


def _sync_user_brief(user: User | None) -> BalanceSyncUser | None:
    if user is None:
        return None
    return BalanceSyncUser(
        id=user.id,
        username=user.username,
        full_name=user.full_name,
        role=user.role,
    )


async def get_sync_info(db: AsyncSession) -> BalanceSyncInfo:
    result = await db.execute(select(SyncMetadata).where(SyncMetadata.id == 1))
    row = result.scalar_one_or_none()
    if row is None or row.last_balances_sync_by is None:
        user = None
    else:
        user = await db.get(User, row.last_balances_sync_by)
    return BalanceSyncInfo(
        last_balances_sync_at=row.last_balances_sync_at if row else None,
        last_balances_sync_by=_sync_user_brief(user),
    )


async def upload_balances(
    db: AsyncSession,
    content: bytes,
    filename: str,
    *,
    user_id: UUID,
) -> BalanceUploadResult:
    lowered = filename.lower()
    if not lowered.endswith((".xlsx", ".xls")):
        raise APIError(400, "INVALID_FILE", "Загрузите файл .xlsx или .xls")
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise APIError(
            400,
            "INVALID_FILE",
            "Не удалось прочитать Excel. Используйте формат .xlsx",
        ) from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise APIError(400, "INVALID_FILE", "Файл пуст")

    error_details: list[BalanceUploadError] = []
    parsed_rows: list[ParsedBalanceRow] = []
    start_index = 1
    if rows and _looks_like_header(_cell(rows[0], EXCEL_COL_PRODUCT)):
        start_index = 2

    for excel_row, row in enumerate(rows[start_index - 1 :], start=start_index):
        if row is None or _row_empty(row):
            continue
        try:
            raw_plant = _cell(row, EXCEL_COL_PLANT)
            raw_warehouse = _cell(row, EXCEL_COL_WAREHOUSE)
            raw_product = _cell(row, EXCEL_COL_PRODUCT)
            product_code = _parse_int_cell(raw_product, "артикул (product_code)")
            erp_plant_code = _parse_int_cell(raw_plant, "завод (erp_plant_code)")
            erp_warehouse_code = _parse_warehouse_code(raw_warehouse)
            logger.info(
                "Excel row %s: plant raw=%r (%s) -> %s; "
                "warehouse raw=%r (%s) -> %r; product raw=%r -> %s",
                excel_row,
                raw_plant,
                type(raw_plant).__name__,
                erp_plant_code,
                raw_warehouse,
                type(raw_warehouse).__name__,
                erp_warehouse_code,
                raw_product,
                product_code,
            )
            available = _parse_quantity(_cell(row, EXCEL_COL_AVAILABLE), "available")
            plan = _parse_quantity(_cell(row, EXCEL_COL_PLAN), "plan")
            unit = _parse_balance_unit(_cell(row, EXCEL_COL_UNIT))
            await _find_plant(db, erp_plant_code)
            warehouse = await _find_warehouse(db, erp_warehouse_code)
            await _find_product(db, product_code)
            parsed_rows.append(
                ParsedBalanceRow(
                    warehouse_code=warehouse.code,
                    product_code=product_code,
                    available=available,
                    plan=plan,
                    unit=unit,
                )
            )
        except (ValueError, APIError) as exc:
            message = exc.message if isinstance(exc, APIError) else str(exc)
            error_details.append(BalanceUploadError(row=excel_row, message=message))

    unique_rows: dict[tuple[int, int], ParsedBalanceRow] = {}
    for parsed in parsed_rows:
        unique_rows[(parsed.warehouse_code, parsed.product_code)] = parsed

    warehouse_codes = {item.warehouse_code for item in unique_rows.values()}
    existing_keys: set[tuple[int, int]] = set()
    if warehouse_codes:
        existing = await db.execute(
            select(
                AvailableBalance.warehouse_code,
                AvailableBalance.product_code,
            ).where(AvailableBalance.warehouse_code.in_(warehouse_codes))
        )
        existing_keys = {(row[0], row[1]) for row in existing.all()}
        await db.execute(
            delete(AvailableBalance).where(
                AvailableBalance.warehouse_code.in_(warehouse_codes)
            )
        )
        await db.flush()

    now = datetime.now(UTC)
    created = 0
    updated = 0
    for key, parsed in unique_rows.items():
        db.add(
            AvailableBalance(
                warehouse_code=parsed.warehouse_code,
                product_code=parsed.product_code,
                available=parsed.available,
                plan=parsed.plan,
                unit=parsed.unit,
                last_sync_at=now,
                source="excel",
            )
        )
        if key in existing_keys:
            updated += 1
        else:
            created += 1

    await _touch_sync_metadata(db, user_id=user_id, synced_at=now)
    await db.commit()
    loaded = created + updated
    return BalanceUploadResult(
        uploaded=loaded,
        created=created,
        updated=updated,
        errors=len(error_details),
        message=f"Загружено {loaded}, ошибок {len(error_details)}",
        error_details=error_details,
    )
