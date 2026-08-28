from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from uuid import UUID

from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select, update
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.exceptions import APIError
from app.core.pagination import paginate
from app.models.normative import Normative
from app.models.object import Object
from app.models.product import Product
from app.models.production_request import ProductionRequest, ProductionRequestItem
from app.models.user import User
from app.schemas.common import PaginationMeta
from app.schemas.production_request import (
    ProductionRequestDatesUpdate,
    ProductionRequestDetail,
    ProductionRequestItemData,
    ProductionRequestListItem,
    ProductionRequestUploadError,
    ProductionRequestUploadOptions,
    ProductionRequestUploadResult,
)
from app.schemas.user import UserBrief

HEADERS = (
    "Завод ERP",
    "Склад ERP",
    "Артикул",
    "Количество",
    "Ед.",
    "Клиент",
)
HEADER_ALIASES = {
    "завод erp": "erp_plant_code",
    "завод_erp": "erp_plant_code",
    "erp_plant_code": "erp_plant_code",
    "склад erp": "erp_warehouse_code",
    "склад_erp": "erp_warehouse_code",
    "erp_warehouse_code": "erp_warehouse_code",
    "артикул": "product_code",
    "product_code": "product_code",
    "количество": "quantity",
    "quantity": "quantity",
    "ед": "unit",
    "ед.": "unit",
    "единица": "unit",
    "unit": "unit",
    "клиент": "client_name",
    "client_name": "client_name",
}
REQUIRED_COLUMNS = {
    "erp_plant_code",
    "erp_warehouse_code",
    "product_code",
    "quantity",
    "unit",
}
ALLOWED_UNITS = {"шт", "кг", "т"}
MAX_QUANTITY = Decimal("9999999999.99")


@dataclass(frozen=True)
class ParsedRow:
    excel_row: int
    erp_plant_code: int
    erp_warehouse_code: str
    product_code: int
    quantity: Decimal
    unit: str
    client_name: str | None


@dataclass(frozen=True)
class ParsedFile:
    rows: list[ParsedRow]
    total_rows: int
    errors: list[ProductionRequestUploadError]


def build_template_xlsx() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Нормативы"
    sheet.append(HEADERS)
    sheet.append([2401, "F005", 10001, 1000, "шт", "ООО Ромашка"])
    sheet.freeze_panes = "A2"
    widths = {"A": 16, "B": 16, "C": 16, "D": 16, "E": 12, "F": 32}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _parse_int(value: object, label: str) -> int:
    if isinstance(value, bool):
        raise ValueError(f"{label}: ожидается целое число")
    try:
        parsed = Decimal(_text(value))
    except InvalidOperation as exc:
        raise ValueError(f"{label}: ожидается целое число") from exc
    if parsed != parsed.to_integral_value():
        raise ValueError(f"{label}: ожидается целое число")
    return int(parsed)


def _parse_erp_warehouse(value: object) -> str:
    text = _text(value).upper()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    if not text:
        raise ValueError("Склад ERP: укажите код")
    if len(text) > 4:
        raise ValueError("Склад ERP должен содержать до 4 символов")
    return text


def _parse_quantity(value: object) -> Decimal:
    try:
        quantity = Decimal(_text(value).replace(",", "."))
    except InvalidOperation as exc:
        raise ValueError("Количество: ожидается число") from exc
    if quantity <= 0:
        raise ValueError("Количество должно быть больше нуля")
    quantity = quantity.quantize(Decimal("0.01"))
    if quantity > MAX_QUANTITY:
        raise ValueError("Количество превышает допустимое значение")
    return quantity


def _parse_unit(value: object) -> str:
    unit = _text(value).lower().replace(" ", "")
    if unit not in ALLOWED_UNITS:
        raise ValueError("Ед.: допустимы шт, кг или т")
    return unit


def _header_map(row: tuple[object, ...]) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, value in enumerate(row):
        normalized = _text(value).lower()
        field = HEADER_ALIASES.get(normalized)
        if field:
            result[field] = index
    missing = REQUIRED_COLUMNS - result.keys()
    if missing:
        labels = ", ".join(sorted(missing))
        raise APIError(
            400,
            "INVALID_FILE",
            f"В Excel отсутствуют обязательные колонки: {labels}",
        )
    expected_order = ("erp_plant_code", "erp_warehouse_code", "product_code")
    if any(result[field] != index for index, field in enumerate(expected_order)):
        raise APIError(
            400,
            "INVALID_FILE",
            "Первые колонки должны идти в порядке: " "Завод ERP, Склад ERP, Артикул",
        )
    return result


def _value(row: tuple[object, ...], columns: dict[str, int], field: str) -> object:
    index = columns.get(field)
    if index is None or index >= len(row):
        return None
    return row[index]


def parse_xlsx(content: bytes, filename: str) -> ParsedFile:
    if not filename.lower().endswith(".xlsx"):
        raise APIError(400, "INVALID_FILE", "Загрузите файл формата .xlsx")
    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise APIError(400, "INVALID_FILE", "Не удалось прочитать Excel") from exc

    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        raise APIError(400, "INVALID_FILE", "Файл пуст")
    columns = _header_map(header)

    parsed: list[ParsedRow] = []
    errors: list[ProductionRequestUploadError] = []
    total_rows = 0
    for row_number, row in enumerate(rows, start=2):
        if not any(_text(value) for value in row):
            continue
        total_rows += 1
        try:
            row_client = _text(_value(row, columns, "client_name")) or None
            parsed.append(
                ParsedRow(
                    excel_row=row_number,
                    erp_plant_code=_parse_int(
                        _value(row, columns, "erp_plant_code"),
                        "Завод ERP",
                    ),
                    erp_warehouse_code=_parse_erp_warehouse(
                        _value(row, columns, "erp_warehouse_code")
                    ),
                    product_code=_parse_int(
                        _value(row, columns, "product_code"),
                        "Артикул",
                    ),
                    quantity=_parse_quantity(_value(row, columns, "quantity")),
                    unit=_parse_unit(_value(row, columns, "unit")),
                    client_name=row_client,
                )
            )
        except ValueError as exc:
            errors.append(
                ProductionRequestUploadError(row=row_number, message=str(exc))
            )
    if total_rows == 0:
        raise APIError(400, "INVALID_FILE", "В файле нет строк с нормативами")
    return ParsedFile(rows=parsed, total_rows=total_rows, errors=errors)


def _to_list_item(batch: ProductionRequest) -> ProductionRequestListItem:
    return ProductionRequestListItem(
        id=batch.id,
        batch_id=batch.batch_id,
        source=batch.source,
        uploaded_by=UserBrief.model_validate(batch.uploader),
        client_name=batch.client_name,
        valid_from=batch.valid_from,
        valid_to=batch.valid_to,
        status=batch.status,
        items_count=len(batch.items),
        created_at=batch.created_at,
    )


def _to_detail(batch: ProductionRequest) -> ProductionRequestDetail:
    return ProductionRequestDetail(
        **_to_list_item(batch).model_dump(),
        items=[
            ProductionRequestItemData(
                id=item.id,
                product_code=item.product_code,
                warehouse_code=item.warehouse_code,
                quantity=item.quantity,
                unit=item.unit,
                client_name=item.client_name,
                category=item.category.strip(),
            )
            for item in batch.items
        ],
    )


async def _load_batch(db: AsyncSession, batch_id: UUID) -> ProductionRequest:
    batch = await db.scalar(
        select(ProductionRequest)
        .options(
            selectinload(ProductionRequest.uploader),
            selectinload(ProductionRequest.items),
        )
        .where(
            ProductionRequest.id == batch_id,
            ProductionRequest.deleted_at.is_(None),
        )
    )
    if batch is None:
        raise APIError(404, "NOT_FOUND", "Партия загрузки не найдена")
    return batch


async def list_batches(
    db: AsyncSession,
    *,
    page: int,
    limit: int,
) -> tuple[list[ProductionRequestListItem], PaginationMeta]:
    conditions = [ProductionRequest.deleted_at.is_(None)]
    total = await db.scalar(
        select(func.count()).select_from(ProductionRequest).where(*conditions)
    )
    result = await db.execute(
        paginate(
            select(ProductionRequest)
            .options(
                selectinload(ProductionRequest.uploader),
                selectinload(ProductionRequest.items),
            )
            .where(*conditions)
            .order_by(ProductionRequest.created_at.desc()),
            page,
            limit,
        )
    )
    return (
        [_to_list_item(batch) for batch in result.scalars().unique().all()],
        PaginationMeta(page=page, limit=limit, total=total or 0),
    )


async def upload_batch(
    db: AsyncSession,
    *,
    content: bytes,
    filename: str,
    user: User,
    options: ProductionRequestUploadOptions,
) -> ProductionRequestUploadResult:
    parsed_file = parse_xlsx(content, filename)
    product_codes = {row.product_code for row in parsed_file.rows}
    erp_plant_codes = {row.erp_plant_code for row in parsed_file.rows}
    products = {
        product.code: product
        for product in (
            await db.scalars(
                select(Product).where(
                    Product.code.in_(product_codes),
                    Product.deleted_at.is_(None),
                    Product.is_active.is_(True),
                )
            )
        ).all()
    }
    objects = (
        await db.scalars(
            select(Object)
            .where(
                Object.erp_plant_code.in_(erp_plant_codes),
                Object.deleted_at.is_(None),
                Object.is_active.is_(True),
            )
            .order_by(Object.code)
        )
    ).all()
    # ERP plant code often lives on warehouses (LogLab 2401) without a
    # separate type=plant row — same rule as logistics balance upload.
    plants = {
        item.erp_plant_code for item in objects if item.erp_plant_code is not None
    }
    warehouses: dict[tuple[int, str], Object] = {}
    for item in objects:
        if (
            item.type == "warehouse"
            and item.erp_plant_code is not None
            and item.erp_warehouse_code
        ):
            warehouses.setdefault(
                (item.erp_plant_code, item.erp_warehouse_code.strip().upper()),
                item,
            )

    errors = list(parsed_file.errors)
    valid_rows: list[tuple[ParsedRow, Product, Object]] = []
    common_client = options.client_name.strip() if options.client_name else None
    for row in parsed_file.rows:
        row_errors: list[str] = []
        product = products.get(row.product_code)
        warehouse = warehouses.get((row.erp_plant_code, row.erp_warehouse_code))
        if row.erp_plant_code not in plants:
            row_errors.append(f"завод ERP {row.erp_plant_code} не найден")
        if row.product_code not in products:
            row_errors.append(f"активный артикул {row.product_code} не найден")
        if warehouse is None:
            row_errors.append(
                "активный склад ERP "
                f"{row.erp_plant_code}/{row.erp_warehouse_code} не найден"
            )
        if not (row.client_name or common_client):
            row_errors.append("укажите клиента в строке или общий клиент")
        if row_errors:
            errors.append(
                ProductionRequestUploadError(
                    row=row.excel_row,
                    message="; ".join(row_errors),
                )
            )
            continue
        if product is not None and warehouse is not None:
            valid_rows.append((row, product, warehouse))

    errors.sort(key=lambda error: error.row)
    if not valid_rows:
        message = f"Загружено 0 строк из {parsed_file.total_rows}"
        return ProductionRequestUploadResult(
            production_request=None,
            total_rows=parsed_file.total_rows,
            imported_count=0,
            error_count=len(errors),
            message=message,
            error_details=errors,
        )

    batch = ProductionRequest(
        uploaded_by=user.id,
        client_name=common_client or None,
        valid_from=options.valid_from,
        valid_to=options.valid_to,
        status="active",
    )
    db.add(batch)
    await db.flush()

    for row, product, warehouse in valid_rows:
        effective_client = row.client_name or common_client or ""
        item = ProductionRequestItem(
            production_request_id=batch.id,
            product_code=row.product_code,
            warehouse_code=warehouse.code,
            quantity=row.quantity,
            unit=row.unit,
            client_name=effective_client,
            category=product.category.strip(),
        )
        db.add(item)
        await db.flush()
        db.add(
            Normative(
                request_id=None,
                production_request_item_id=item.id,
                product_code=row.product_code,
                warehouse_code=warehouse.code,
                quantity=row.quantity,
                unit=row.unit,
                client_name=effective_client,
                expiry_date=options.valid_to,
                category=product.category.strip(),
            )
        )

    await db.commit()
    loaded = await _load_batch(db, batch.id)
    message = f"Загружено {len(valid_rows)} строк из {parsed_file.total_rows}"
    return ProductionRequestUploadResult(
        production_request=_to_detail(loaded),
        total_rows=parsed_file.total_rows,
        imported_count=len(valid_rows),
        error_count=len(errors),
        message=message,
        error_details=errors,
    )


async def update_dates(
    db: AsyncSession,
    *,
    batch_id: UUID,
    body: ProductionRequestDatesUpdate,
) -> ProductionRequestDetail:
    batch = await _load_batch(db, batch_id)
    batch.valid_from = body.valid_from
    batch.valid_to = body.valid_to
    item_ids = select(ProductionRequestItem.id).where(
        ProductionRequestItem.production_request_id == batch.id
    )
    await db.execute(
        update(Normative)
        .where(Normative.production_request_item_id.in_(item_ids))
        .values(expiry_date=body.valid_to)
    )
    await db.commit()
    return _to_detail(await _load_batch(db, batch.id))


async def delete_batch(db: AsyncSession, *, batch_id: UUID) -> None:
    batch = await _load_batch(db, batch_id)
    await db.delete(batch)
    await db.commit()
